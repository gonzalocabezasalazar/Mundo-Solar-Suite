"""
pages/global_view.py
══════════════════════════════════════════════════════════════
Vista Global — todas las plantas con KPIs consolidados.
Optimizado con Vectorización Aislada (Evita mezcla de cajas homónimas).
══════════════════════════════════════════════════════════════
"""
import streamlit as st
import pandas as pd
import plotly.express as px
from datetime import timedelta

from components.filters import flexible_period_filter
from components.cards import planta_card, kpi_row
from components.theme import get_colors, theme_toggle_button
from ms_data.analysis import analizar_mediciones, _to_float

def render(df_plantas, df_fallas, df_med, df_tec):
    c = get_colors()
    hoy = pd.Timestamp.now()

    # ── Header ──────────────────────────────────────────────
    col_logo, col_toggle = st.columns([5, 1])
    with col_logo:
        st.markdown("""
        <div class="suite-logo">
            <div class="logo-icon">☀️</div>
            <div class="logo-text">
                <h1>Mundo Solar Suite</h1>
                <p>Panel de control global — O&M de plantas PMGD</p>
            </div>
            <div class="logo-badge">LIVE</div>
        </div>
        """, unsafe_allow_html=True)
    with col_toggle:
        st.markdown("<br>", unsafe_allow_html=True)
        theme_toggle_button()

    if df_plantas.empty:
        st.warning("No hay plantas registradas. Ve a ⚙️ Gestión Plantas para agregar.")
        return

    # ── Selector de período (Popover) ───────────────────────
    fil = flexible_period_filter(
        key='vg_main',
        df_med=df_med,
        df_fallas=df_fallas,
        default_mode='Mes'
    )
    m_fil = fil['df_med']
    f_fil = fil['df_fallas']
    lbl   = fil['label']

    st.divider()

    # ── KPIs globales ────────────────────────────────────────
    total_plantas = len(df_plantas)
    total_fallas  = len(f_fil) if f_fil is not None and not f_fil.empty else 0
    total_meds    = len(m_fil) if m_fil is not None and not m_fil.empty else 0
    total_tecs    = len(df_tec[df_tec.get('Activo', 'SI') == 'SI']) if not df_tec.empty else 0
    
    fallas_mes = 0
    if not df_fallas.empty and 'Fecha' in df_fallas.columns:
        df_fallas['Fecha'] = pd.to_datetime(df_fallas['Fecha'], errors='coerce')
        fallas_mes = len(df_fallas[df_fallas['Fecha'].dt.month == hoy.month])

    kpi_row([
        {'label': 'Plantas Activas',  'value': total_plantas, 'cls': 'gold'},
        {'label': f'Fallas ({lbl})',  'value': total_fallas,  'cls': 'crit' if total_fallas > 0 else 'ok'},
        {'label': 'Fallas este mes',  'value': fallas_mes,    'cls': 'warn' if fallas_mes > 0 else 'ok'},
        {'label': 'Mediciones',       'value': total_meds,    'cls': ''},
        {'label': 'Técnicos O&M',     'value': total_tecs,    'cls': ''},
    ])

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="section-hdr">📍 Estado Operativo por Planta</div>', unsafe_allow_html=True)

    # ── OPTIMIZACIÓN CORREGIDA: Calcular salud aislando por planta ──
    vg_fallas = f_fil.groupby('Planta_ID').size().to_dict() if f_fil is not None and not f_fil.empty else {}
    
    dict_salud = {}
    if m_fil is not None and not m_fil.empty:
        # El groupby evita la contaminación cruzada de Cajas (Ej: CB-1 de Planta A con CB-1 de Planta B)
        for pid, df_p in m_fil.groupby('Planta_ID'):
            df_an = analizar_mediciones(df_p)
            total_str = len(df_an)
            n_norm = len(df_an[df_an['Diagnostico'] == 'NORMAL'])
            n_crit = len(df_an[df_an['Diagnostico'].isin(['CRÍTICO', 'OC (0A)'])])
            n_aler = len(df_an[df_an['Diagnostico'] == 'ALERTA'])
            salud_pct = (n_norm / total_str * 100) if total_str > 0 else 100.0
            dict_salud[pid] = {'salud': salud_pct, 'crit': n_crit, 'aler': n_aler}

    # ── Tarjetas por planta ──────────────────────────────────
    n_cols = min(3, len(df_plantas))
    cols = st.columns(n_cols)

    for i, (_, planta) in enumerate(df_plantas.iterrows()):
        pid        = str(planta['ID'])
        nombre     = planta.get('Nombre', f'Planta {pid}')
        ubicacion  = planta.get('Ubicacion', '')
        tecnologia = planta.get('Tecnologia', '')
        potencia   = _to_float(planta.get('Potencia_MW', 0))

        stats = dict_salud.get(pid, {'salud': 100.0, 'crit': 0, 'aler': 0})
        n_fallas = vg_fallas.get(pid, 0)

        # Calculo rápido de salud del mes anterior (Delta)
        salud_ant = None
        if not df_med.empty:
            mes_ant = (hoy - pd.DateOffset(months=1)).to_period('M')
            m_ant = df_med[(df_med['Planta_ID'] == pid) & (df_med['Fecha'].dt.to_period('M') == mes_ant)]
            if not m_ant.empty:
                df_an_ant = analizar_mediciones(m_ant)
                if not df_an_ant.empty:
                    salud_ant = (len(df_an_ant[df_an_ant['Diagnostico'] == 'NORMAL']) / len(df_an_ant)) * 100

        with cols[i % n_cols]:
            planta_card(
                planta_id       = pid,
                nombre          = nombre,
                ubicacion       = ubicacion,
                tecnologia      = tecnologia,
                potencia_mw     = potencia,
                salud_pct       = stats['salud'],
                salud_anterior_pct = salud_ant,
                n_fallas        = n_fallas,
                n_criticos      = stats['crit'],
                n_alertas       = stats['aler'],
            )

    # ── Gráficos consolidados ────────────────────────────────
    st.markdown('<div class="section-hdr">📊 Análisis Consolidado Global</div>', unsafe_allow_html=True)
    c_colors = get_colors()
    col1, col2 = st.columns(2)

    nombres_validos = df_plantas['Nombre'].astype(str).str.strip().tolist()

    with col1:
        if f_fil is not None and not f_fil.empty and 'Planta_Nombre' in f_fil.columns:
            df_f_ok = f_fil[f_fil['Planta_Nombre'].astype(str).str.strip().isin(nombres_validos)]
            if not df_f_ok.empty:
                df_fxp = df_f_ok.groupby('Planta_Nombre').size().reset_index(name='Fallas')
                fig = px.bar(df_fxp, x='Planta_Nombre', y='Fallas',
                             color='Fallas', color_continuous_scale='Reds',
                             title=f"Fusibles operados — {lbl}", text='Fallas')
                fig.update_traces(textposition='outside')
                fig.update_layout(height=320, showlegend=False,
                                  plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                                  xaxis_title='', yaxis_title='N° Fusibles',
                                  font_color=c_colors['text'], margin=dict(t=40, b=30, l=20, r=20))
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("Sin fallas en el período seleccionado para las plantas activas.")
        else:
            st.info("Sin fallas registradas en el período.")

    with col2:
        if m_fil is not None and not m_fil.empty and 'Planta_Nombre' in m_fil.columns:
            df_m_ok = m_fil[m_fil['Planta_Nombre'].astype(str).str.strip().isin(nombres_validos)]
            if not df_m_ok.empty:
                df_mxp = df_m_ok.groupby('Planta_Nombre')['Amperios'].mean().reset_index()
                df_mxp.columns = ['Planta', 'I Media (A)']
                df_mxp['I Media (A)'] = df_mxp['I Media (A)'].round(3)
                
                _imin, _imax = df_mxp['I Media (A)'].min(), df_mxp['I Media (A)'].max()
                _margen = max((_imax - _imin) * 0.4, 0.3)
                _ymin, _ymax = max(0, _imin - _margen), _imax + _margen

                fig2 = px.bar(df_mxp, x='Planta', y='I Media (A)',
                              color='I Media (A)', color_continuous_scale='Blues',
                              range_color=[0, _ymax],
                              title=f"Corriente media global — {lbl}", text='I Media (A)')
                fig2.update_traces(textposition='outside')
                fig2.update_layout(height=320, showlegend=False,
                                   plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                                   xaxis_title='', yaxis_title='I Media (A)',
                                   yaxis=dict(range=[_ymin, _ymax]),
                                   font_color=c_colors['text'], margin=dict(t=40, b=30, l=20, r=20))
                st.plotly_chart(fig2, use_container_width=True)
            else:
                st.info("Sin mediciones para las plantas activas.")
        else:
            st.info("Sin mediciones registradas en el período.")

    # ── Alertas recientes ────────────────────────────────────
    if not df_fallas.empty and 'Fecha' in df_fallas.columns:
        recientes = df_fallas[df_fallas['Fecha'] >= (hoy - timedelta(days=30))]
        if not recientes.empty:
            st.markdown('<div class="section-hdr">🔔 Últimas Fallas (30 días)</div>', unsafe_allow_html=True)
            cols_show = [c for c in ['Fecha', 'Planta_Nombre', 'Inversor', 'Caja', 'String', 'Amperios', 'Nota'] if c in recientes.columns]
            df_show = recientes[cols_show].copy()
            df_show['Fecha'] = df_show['Fecha'].dt.strftime('%d/%m/%Y')
            st.dataframe(df_show.sort_values('Fecha', ascending=False).head(15), use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════
# VISTA LECTOR (Simplificada)
# ══════════════════════════════════════════════════════════════
def render_kpis(df_plantas, df_fallas, df_med):
    from ms_data.exports import generar_excel_mediciones
    import io
    from openpyxl import Workbook
    from ms_data.analysis import _run_in_thread

    c = get_colors()
    hoy = pd.Timestamp.now()
    usr = st.session_state.get('usuario', {})

    col_logo, col_toggle = st.columns([5, 1])
    with col_logo:
        st.markdown("""
        <div class="suite-logo">
            <div class="logo-icon">📊</div>
            <div class="logo-text">
                <h1>Panel de KPIs</h1>
                <p>Resumen operacional — Modo Lectura</p>
            </div>
        </div>
        """, unsafe_allow_html=True)
    with col_toggle:
        st.markdown("<br>", unsafe_allow_html=True)
        theme_toggle_button()

    st.caption(f"Sesión: {usr.get('nombre','')} · {hoy.strftime('%d/%m/%Y')}")

    fil = flexible_period_filter(key='kpis_lector', df_med=df_med, df_fallas=df_fallas, default_mode="Mes")
    m_fil, f_fil, lbl = fil['df_med'], fil['df_fallas'], fil['label']
    st.caption(f"Mostrando datos para: **{lbl}**")
    st.divider()

    if df_plantas.empty:
        st.info("No hay plantas registradas.")
        return

    # Aislamos las métricas por planta también para el modo Lector
    dict_salud = {}
    if m_fil is not None and not m_fil.empty:
        for pid, df_p in m_fil.groupby('Planta_ID'):
            dan = analizar_mediciones(df_p)
            t_str = len(dan)
            n_nrm = len(dan[dan['Diagnostico'] == 'NORMAL'])
            n_crt = len(dan[dan['Diagnostico'].isin(['CRÍTICO', 'OC (0A)'])])
            dict_salud[pid] = {'salud': (n_nrm / t_str * 100) if t_str > 0 else 100.0, 'crit': n_crt}

    datos = []
    for _, planta in df_plantas.iterrows():
        pid = str(planta['ID'])
        stats = dict_salud.get(pid, {'salud': 100.0, 'crit': 0})
        fp = f_fil[f_fil['Planta_ID'] == pid] if f_fil is not None and not f_fil.empty else []
        mp = m_fil[m_fil['Planta_ID'] == pid] if m_fil is not None and not m_fil.empty else []
        
        datos.append({
            'nombre': planta.get('Nombre', pid),
            'salud': stats['salud'], 'n_crit': stats['crit'],
            'n_fallas': len(fp), 'n_meds': len(mp),
            'i_prom': mp['Amperios'].mean() if len(mp) > 0 else 0,
        })

    salud_global = sum(d['salud'] for d in datos) / len(datos) if datos else 100
    criticos_total = sum(d['n_crit'] for d in datos)
    fallas_total = sum(d['n_fallas'] for d in datos)
    sem = '🟢' if salud_global >= 90 else '🟡' if salud_global >= 70 else '🔴'

    kpi_row([
        {'label': 'Plantas',        'value': len(datos),          'cls': 'gold'},
        {'label': f'{sem} Salud',   'value': f'{salud_global:.1f}%', 'cls': 'ok' if salud_global >= 90 else 'warn' if salud_global >= 70 else 'crit'},
        {'label': 'Fallas',         'value': fallas_total,        'cls': 'crit' if fallas_total > 0 else 'ok'},
        {'label': 'Críticos',       'value': criticos_total,      'cls': 'crit' if criticos_total > 0 else 'ok'},
    ])

    st.divider()
    st.markdown('<div class="section-hdr">🏭 Estado por Planta</div>', unsafe_allow_html=True)

    filas = []
    for d in datos:
        sem_p = '🟢' if d['salud'] >= 90 else '🟡' if d['salud'] >= 70 else '🔴'
        filas.append({
            'Planta': f"{sem_p} {d['nombre']}", 'Salud %': round(d['salud'], 1),
            'Fallas': d['n_fallas'], 'Mediciones': d['n_meds'],
            'I Prom (A)': round(d['i_prom'], 3) if d['i_prom'] else '—',
            'Críticos': d['n_crit'],
        })
    df_res = pd.DataFrame(filas)
    st.dataframe(df_res, use_container_width=True, hide_index=True)

    if df_med is not None and not df_med.empty:
        tend = []
        df_tmp = df_med.copy()
        df_tmp['Mes'] = df_tmp['Fecha'].dt.strftime('%Y-%m')
        for mes in sorted(df_tmp['Mes'].unique()):
            saludes_mes = []
            for _, planta in df_plantas.iterrows():
                pid = str(planta['ID'])
                mp = df_tmp[(df_tmp['Mes'] == mes) & (df_tmp['Planta_ID'] == pid)]
                if mp.empty: continue
                dan = analizar_mediciones(mp)
                if not dan.empty: saludes_mes.append(len(dan[dan['Diagnostico'] == 'NORMAL']) / len(dan) * 100)
            if saludes_mes: tend.append({'Mes': mes, 'Salud %': round(sum(saludes_mes)/len(saludes_mes), 1)})
        if tend:
            df_tend = pd.DataFrame(tend)
            fig = px.line(df_tend, x='Mes', y='Salud %', markers=True, title='Tendencia Salud Global', color_discrete_sequence=[c['ok']])
            fig.add_hline(y=90, line_dash='dash', line_color=c['warn'], annotation_text='Meta 90%')
            fig.update_layout(height=280, plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', font_color=c['text'], yaxis=dict(range=[0, 105]), xaxis_title='', margin=dict(t=40, b=20))
            st.plotly_chart(fig, use_container_width=True)

    st.divider()
    def _gen_excel():
        wb = Workbook(); ws = wb.active; ws.title = 'KPIs'
        ws.merge_cells('A1:F1')
        ws['A1'] = f'Reporte KPIs Mundo Solar — {lbl}'
        ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
        ws['A1'].fill = PatternFill('solid', fgColor='1A3A5C')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.append([f'Generado: {hoy.strftime("%d/%m/%Y %H:%M")}']); ws.append([])
        ws.append(list(df_res.columns))
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True, color='FFFFFF'); cell.fill = PatternFill('solid', fgColor='2E86AB'); cell.alignment = Alignment(horizontal='center')
        for row in df_res.values: ws.append(list(row))
        for i in range(1, len(df_res.columns) + 1): ws.column_dimensions[get_column_letter(i)].width = 22
        out = io.BytesIO(); wb.save(out); out.seek(0)
        return out.getvalue()

    xls = _run_in_thread(_gen_excel)
    st.download_button('📊 Descargar Excel KPIs', xls, f'KPIs_MundoSolar_{hoy.strftime("%Y%m")}.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', use_container_width=True)