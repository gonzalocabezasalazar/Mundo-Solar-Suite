"""
pages/global_view.py
══════════════════════════════════════════════════════════════
Vista Global — todas las plantas con KPIs consolidados.
Accesible por admin y técnico.
Vista simplificada (kpis_view) para rol lector.
══════════════════════════════════════════════════════════════
"""
import streamlit as st
import pandas as pd
import plotly.express as px
from datetime import timedelta

from components.filters import period_selector
from components.cards import planta_card, kpi_row, breadcrumb
from components.theme import get_colors, theme_toggle_button
from ms_data.analysis import analizar_mediciones


# ══════════════════════════════════════════════════════════════
# VISTA GLOBAL — admin / técnico
# ══════════════════════════════════════════════════════════════
def render(df_plantas, df_fallas, df_med, df_tec):
    c = get_colors()
    hoy = pd.Timestamp.now()

    # ── Header ──────────────────────────────────────────────
    col_logo, col_toggle = st.columns([5, 1])
    with col_logo:
        st.markdown("""
        <div class="suite-logo">
            <div class="logo-icon">☀️</div>
            <div class="logo-text">
                <h1>Mundo Solar Suite</h1>
                <p>Panel de control global — todas las plantas PMGD</p>
            </div>
            <div class="logo-badge">LIVE</div>
        </div>
        """, unsafe_allow_html=True)
    with col_toggle:
        st.markdown("<br>", unsafe_allow_html=True)
        theme_toggle_button()

    if df_plantas.empty:
        st.warning("No hay plantas registradas. Ve a ⚙️ Gestión Plantas para agregar.")
        return

    # ── Selector de período ──────────────────────────────────
    col_per, _ = st.columns([1, 3])
    with col_per:
        fil = period_selector(
            key='vg',
            df_med=df_med,
            df_fallas=df_fallas,
        )
    m_fil = fil['df_med']
    f_fil = fil['df_fallas']
    lbl   = fil['label']

    st.caption(f"📅 Mostrando: **{lbl}**")

    # ── KPIs globales ────────────────────────────────────────
    total_plantas = len(df_plantas)
    total_fallas  = len(f_fil) if f_fil is not None and not f_fil.empty else 0
    total_meds    = len(m_fil) if m_fil is not None and not m_fil.empty else 0
    total_tecs    = len(df_tec[df_tec.get('Activo', 'SI') == 'SI']) if not df_tec.empty else 0
    fallas_mes    = 0
    if not df_fallas.empty and 'Fecha' in df_fallas.columns:
        fallas_mes = len(df_fallas[df_fallas['Fecha'].dt.month == hoy.month])

    kpi_row([
        {'label': 'Plantas Activas',  'value': total_plantas, 'cls': 'gold'},
        {'label': 'Fallas período',   'value': total_fallas,
         'cls': 'crit' if total_fallas > 0 else 'ok'},
        {'label': 'Fallas este mes',  'value': fallas_mes,
         'cls': 'warn' if fallas_mes > 0 else 'ok'},
        {'label': 'Mediciones DB',    'value': total_meds,    'cls': ''},
        {'label': 'Técnicos activos', 'value': total_tecs,    'cls': ''},
    ])

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="section-hdr">📍 Estado por Planta</div>',
                unsafe_allow_html=True)

    # ── Precalcular métricas por planta (vectorizado) ────────
    vg_fallas = f_fil.groupby('Planta_ID').size().to_dict() \
                if f_fil is not None and not f_fil.empty else {}
    vg_meds   = m_fil.groupby('Planta_ID').size().to_dict() \
                if m_fil is not None and not m_fil.empty else {}

    # Calcular salud y delta por planta
    def _salud_planta(pid, df_m):
        mp = df_m[df_m['Planta_ID'] == pid] if not df_m.empty else pd.DataFrame()
        if mp.empty:
            return 100.0, 0, 0
        dan = analizar_mediciones(mp)
        if dan.empty:
            return 100.0, 0, 0
        s = len(dan[dan['Diagnostico'] == 'NORMAL']) / len(dan) * 100
        n_crit = len(dan[dan['Diagnostico'].isin(['CRÍTICO', 'OC (0A)'])])
        n_aler = len(dan[dan['Diagnostico'] == 'ALERTA'])
        return s, n_crit, n_aler

    def _salud_mes_anterior(pid, df_m_full):
        """Salud del mes anterior para calcular delta."""
        if df_m_full.empty:
            return None
        mes_ant = (hoy - pd.DateOffset(months=1)).to_period('M')
        mp = df_m_full[
            (df_m_full['Planta_ID'] == pid) &
            (df_m_full['Fecha'].dt.to_period('M') == mes_ant)
        ]
        if mp.empty:
            return None
        dan = analizar_mediciones(mp)
        if dan.empty:
            return None
        return len(dan[dan['Diagnostico'] == 'NORMAL']) / len(dan) * 100

    # ── Tarjetas por planta ──────────────────────────────────
    n_cols = min(3, len(df_plantas))
    cols = st.columns(n_cols)

    for i, (_, planta) in enumerate(df_plantas.iterrows()):
        pid       = str(planta['ID'])
        nombre    = planta.get('Nombre', f'Planta {pid}')
        ubicacion = planta.get('Ubicacion', '')
        tecnologia= planta.get('Tecnologia', '')
        potencia  = planta.get('Potencia_MW', 0)

        m_p = m_fil[m_fil['Planta_ID'] == pid] \
              if m_fil is not None and not m_fil.empty else pd.DataFrame()

        salud, n_crit, n_aler = _salud_planta(pid, m_fil if m_fil is not None else pd.DataFrame())
        salud_ant = _salud_mes_anterior(pid, df_med)
        n_fallas  = vg_fallas.get(pid, 0)

        with cols[i % n_cols]:
            planta_card(
                planta_id       = pid,
                nombre          = nombre,
                ubicacion       = ubicacion,
                tecnologia      = tecnologia,
                potencia_mw     = potencia,
                salud_pct       = salud,
                salud_anterior_pct = salud_ant,
                n_fallas        = n_fallas,
                n_criticos      = n_crit,
                n_alertas       = n_aler,
            )

    # ── Gráficos consolidados ────────────────────────────────
    if f_fil is not None and not f_fil.empty and 'Planta_Nombre' in f_fil.columns:
        st.markdown('<div class="section-hdr">📊 Análisis Consolidado</div>',
                    unsafe_allow_html=True)

        nombres_validos = df_plantas['Nombre'].astype(str).str.strip().tolist()
        df_f_ok = f_fil[f_fil['Planta_Nombre'].astype(str).str.strip().isin(nombres_validos)]
        df_m_ok = m_fil[m_fil['Planta_Nombre'].astype(str).str.strip().isin(nombres_validos)] \
                  if m_fil is not None and not m_fil.empty and 'Planta_Nombre' in m_fil.columns \
                  else pd.DataFrame()

        c_colors = get_colors()
        col1, col2 = st.columns(2)

        with col1:
            df_fxp = df_f_ok.groupby('Planta_Nombre').size().reset_index(name='Fallas')
            fig = px.bar(df_fxp, x='Planta_Nombre', y='Fallas',
                         color='Fallas', color_continuous_scale='Reds',
                         title=f"Fusibles quemados — {lbl}", text='Fallas')
            fig.update_traces(textposition='outside')
            fig.update_layout(height=320, showlegend=False,
                              plot_bgcolor='rgba(0,0,0,0)',
                              paper_bgcolor='rgba(0,0,0,0)',
                              xaxis_title='', yaxis_title='N° Fusibles',
                              font_color=c_colors['text'],
                              margin=dict(t=40, b=30, l=20, r=20))
            st.plotly_chart(fig, use_container_width=True)

        with col2:
            if not df_m_ok.empty:
                df_mxp = df_m_ok.groupby('Planta_Nombre')['Amperios'].mean().reset_index()
                df_mxp.columns = ['Planta', 'I Media (A)']
                df_mxp['I Media (A)'] = df_mxp['I Media (A)'].round(3)
                # Eje Y dinámico: empieza cerca del mínimo para que barras sean visibles
                _imin = df_mxp['I Media (A)'].min()
                _imax = df_mxp['I Media (A)'].max()
                _margen = max((_imax - _imin) * 0.4, 0.3)
                _ymin = max(0, _imin - _margen)
                _ymax = _imax + _margen

                fig2 = px.bar(df_mxp, x='Planta', y='I Media (A)',
                              color='I Media (A)', color_continuous_scale='Blues',
                              title=f"Corriente media — {lbl}", text='I Media (A)')
                fig2.update_traces(textposition='outside')
                fig2.update_layout(height=320, showlegend=False,
                                   plot_bgcolor='rgba(0,0,0,0)',
                                   paper_bgcolor='rgba(0,0,0,0)',
                                   xaxis_title='', yaxis_title='I Media (A)',
                                   yaxis=dict(range=[_ymin, _ymax]),
                                   font_color=c_colors['text'],
                                   margin=dict(t=40, b=30, l=20, r=20))
                st.plotly_chart(fig2, use_container_width=True)

    # ── Alertas recientes ────────────────────────────────────
    if not df_fallas.empty and 'Fecha' in df_fallas.columns:
        recientes = df_fallas[df_fallas['Fecha'] >= (hoy - timedelta(days=30))]
        if not recientes.empty:
            st.markdown(
                '<div class="section-hdr">🔔 Alertas Recientes (últimos 30 días)</div>',
                unsafe_allow_html=True)
            cols_show = [c for c in
                         ['Fecha', 'Planta_Nombre', 'Inversor', 'Caja',
                          'String', 'Amperios', 'Nota']
                         if c in recientes.columns]
            df_show = recientes[cols_show].copy()
            df_show['Fecha'] = df_show['Fecha'].dt.strftime('%d/%m/%Y')
            st.dataframe(
                df_show.sort_values('Fecha', ascending=False).head(20),
                use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════
# VISTA KPIs — solo lectura (rol lector)
# ══════════════════════════════════════════════════════════════
def render_kpis(df_plantas, df_fallas, df_med):
    """Vista simplificada para rol lector — KPIs + descarga."""
    from ms_data.analysis import _run_in_thread
    from ms_data.exports import generar_excel_mediciones
    import io
    from openpyxl import Workbook

    c = get_colors()
    hoy = pd.Timestamp.now()
    usr = st.session_state.get('usuario', {})

    col_logo, col_toggle = st.columns([5, 1])
    with col_logo:
        st.markdown("""
        <div class="suite-logo">
            <div class="logo-icon">📊</div>
            <div class="logo-text">
                <h1>Panel de KPIs</h1>
                <p>Resumen operacional — todas las plantas</p>
            </div>
        </div>
        """, unsafe_allow_html=True)
    with col_toggle:
        st.markdown("<br>", unsafe_allow_html=True)
        theme_toggle_button()

    st.caption(f"Vista de solo lectura · {usr.get('nombre','')} · "
               f"{hoy.strftime('%d/%m/%Y %H:%M')}")

    # ── Selector período ──────────────────────────────────────
    col_per, _ = st.columns([1, 3])
    with col_per:
        fil = period_selector(key='kpis', df_med=df_med, df_fallas=df_fallas)
    m_fil  = fil['df_med']
    f_fil  = fil['df_fallas']
    lbl    = fil['label']
    st.caption(f"📅 **{lbl}**")
    st.divider()

    if df_plantas.empty:
        st.info("No hay plantas registradas.")
        return

    # ── KPIs globales ─────────────────────────────────────────
    datos = []
    for _, planta in df_plantas.iterrows():
        pid  = str(planta['ID'])
        mp   = m_fil[m_fil['Planta_ID'] == pid] \
               if m_fil is not None and not m_fil.empty else pd.DataFrame()
        fp   = f_fil[f_fil['Planta_ID'] == pid] \
               if f_fil is not None and not f_fil.empty else pd.DataFrame()

        if not mp.empty:
            dan    = analizar_mediciones(mp)
            salud  = len(dan[dan['Diagnostico'] == 'NORMAL']) / len(dan) * 100
            n_crit = len(dan[dan['Diagnostico'].isin(['CRÍTICO', 'OC (0A)'])])
        else:
            salud, n_crit = 100.0, 0

        datos.append({
            'pid': pid, 'nombre': planta.get('Nombre', pid),
            'salud': salud, 'n_crit': n_crit,
            'n_fallas': len(fp), 'n_meds': len(mp),
            'i_prom': mp['Amperios'].mean() if not mp.empty else 0,
        })

    salud_global   = sum(d['salud'] for d in datos) / len(datos) if datos else 100
    criticos_total = sum(d['n_crit'] for d in datos)
    fallas_total   = sum(d['n_fallas'] for d in datos)
    sem = '🟢' if salud_global >= 90 else '🟡' if salud_global >= 70 else '🔴'

    kpi_row([
        {'label': 'Plantas',        'value': len(datos),          'cls': 'gold'},
        {'label': f'{sem} Salud Global', 'value': f'{salud_global:.1f}%',
         'cls': 'ok' if salud_global >= 90 else 'warn' if salud_global >= 70 else 'crit'},
        {'label': 'Fallas',         'value': fallas_total,
         'cls': 'crit' if fallas_total > 0 else 'ok'},
        {'label': 'Strings Críticos','value': criticos_total,
         'cls': 'crit' if criticos_total > 0 else 'ok'},
    ])

    # ── Tabla resumen ─────────────────────────────────────────
    st.divider()
    st.markdown('<div class="section-hdr">🏭 Estado por Planta</div>',
                unsafe_allow_html=True)

    filas = []
    for d in datos:
        sem_p = '🟢' if d['salud'] >= 90 else '🟡' if d['salud'] >= 70 else '🔴'
        filas.append({
            'Planta':            f"{sem_p} {d['nombre']}",
            'Salud %':           round(d['salud'], 1),
            'Fallas':            d['n_fallas'],
            'Mediciones':        d['n_meds'],
            'I Prom (A)':        round(d['i_prom'], 3) if d['i_prom'] else '—',
            'Strings Críticos':  d['n_crit'],
        })
    df_res = pd.DataFrame(filas)
    st.dataframe(df_res, use_container_width=True, hide_index=True)

    # ── Gráfico tendencia salud ───────────────────────────────
    if df_med is not None and not df_med.empty:
        tend = []
        df_tmp = df_med.copy()
        df_tmp['Mes'] = df_tmp['Fecha'].dt.strftime('%Y-%m')
        for mes in sorted(df_tmp['Mes'].unique()):
            saludes_mes = []
            for _, planta in df_plantas.iterrows():
                pid = str(planta['ID'])
                mp  = df_tmp[(df_tmp['Mes'] == mes) & (df_tmp['Planta_ID'] == pid)]
                if mp.empty:
                    continue
                dan = analizar_mediciones(mp)
                if not dan.empty:
                    saludes_mes.append(
                        len(dan[dan['Diagnostico'] == 'NORMAL']) / len(dan) * 100)
            if saludes_mes:
                tend.append({'Mes': mes,
                             'Salud %': round(sum(saludes_mes)/len(saludes_mes), 1)})
        if tend:
            df_tend = pd.DataFrame(tend)
            fig = px.line(df_tend, x='Mes', y='Salud %', markers=True,
                          title='Tendencia de Salud Global',
                          color_discrete_sequence=[c['ok']])
            fig.add_hline(y=90, line_dash='dash', line_color=c['warn'],
                          annotation_text='Meta 90%')
            fig.update_layout(height=280,
                              plot_bgcolor='rgba(0,0,0,0)',
                              paper_bgcolor='rgba(0,0,0,0)',
                              font_color=c['text'],
                              yaxis=dict(range=[0, 105]),
                              xaxis_title='', margin=dict(t=40, b=20))
            fig.update_xaxes(tickangle=-35, type='category')
            st.plotly_chart(fig, use_container_width=True)

    # ── Descarga Excel ────────────────────────────────────────
    st.divider()
    def _gen_excel():
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = 'KPIs'
        ws.merge_cells('A1:F1')
        ws['A1'] = f'Reporte KPIs Mundo Solar — {lbl}'
        ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
        ws['A1'].fill = PatternFill('solid', fgColor='1A3A5C')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.append([f'Generado: {hoy.strftime("%d/%m/%Y %H:%M")}'])
        ws.append([])
        ws.append(list(df_res.columns))
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='2E86AB')
            cell.alignment = Alignment(horizontal='center')
        for row in df_res.values:
            ws.append(list(row))
        for i in range(1, len(df_res.columns) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 22
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()

    xls = _run_in_thread(_gen_excel)
    st.download_button(
        '📊 Descargar Excel KPIs', xls,
        f'KPIs_MundoSolar_{hoy.strftime("%Y%m")}.xlsx',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        use_container_width=True)
