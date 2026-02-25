"""
data/exports.py
══════════════════════════════════════════════════════════════
Generación de reportes PDF y Excel.
Sin lógica de UI — funciones puras que retornan bytes.
Siempre llamar via _run_in_thread() para evitar conflictos con Streamlit.
══════════════════════════════════════════════════════════════
"""
import io
import datetime
import tempfile
import os

import pandas as pd
import numpy as np
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Dependencias internas — TODAS las utilidades de análisis necesarias
from ms_data.analysis import (
    clean_text, _to_float, _to_int, obtener_nombre_mes, 
    clasificar_falla_amp, clasificar_falla_isc, analizar_mediciones
)
# ══════════════════════════════════════════════════════════════
# PDF ENGINE
# ══════════════════════════════════════════════════════════════
class PDF(FPDF):
    def header(self):
        self.set_font('Arial','B',8)
        self.set_text_color(120,120,120)
        self.cell(0,5,'Elaborado por Mundo Solar SpA',0,1,'R')
        self.set_font('Arial','B',13)
        self.set_text_color(26,58,92)
        self.cell(0,10,'INFORME TECNICO PMGD - MUNDO SOLAR SUITE',0,1,'C')
        self.ln(3)
    def footer(self):
        self.set_y(-14)
        self.set_font('Arial','I',8)
        self.set_text_color(150,150,150)
        self.cell(0,10,f'Pagina {self.page_no()} - Mundo Solar Suite',0,0,'C')

def generar_pdf_fallas(planta_nombre, df_fallas, df_med=None, cfg=None):
    """PDF de fallas enriquecido con datos de mediciones de strings."""
    pdf = PDF()
    pdf.add_page()

    # Calcular Tipo con ajuste climático
    isc_stc = _to_float(cfg.get('Isc_STC_A', 9.07)) if cfg else 9.07
    df_fallas = df_fallas.copy()
    def _tipo_row(r):
        irr = _to_float(r.get('Irradiancia_Wm2', 0))
        return clasificar_falla_isc(r['Amperios'], isc_stc, irr) if irr > 50 else clasificar_falla_amp(r['Amperios'])
    df_fallas['Tipo'] = df_fallas.apply(_tipo_row, axis=1)

    # KPIs desde mediciones si están disponibles
    prom_med   = df_med['Amperios'].mean() if df_med is not None and not df_med.empty else None
    total_str  = len(df_med) if df_med is not None and not df_med.empty else 0

    # Strings con desviación desde mediciones
    df_anom_med = pd.DataFrame()
    if df_med is not None and not df_med.empty and 'Desv_CB_pct' in df_med.columns:
        df_anom_med = df_med[df_med['Diagnostico'] != 'NORMAL'].copy()
    elif df_med is not None and not df_med.empty:
        df_med2 = df_med.copy()
        df_med2['Promedio_Caja'] = df_med2.groupby('Equipo')['Amperios'].transform('mean')
        df_med2['Desv_CB_pct']  = np.where(
            df_med2['Promedio_Caja'] > 0,
            ((df_med2['Amperios'] - df_med2['Promedio_Caja']) / df_med2['Promedio_Caja']) * 100, 0)
        ua = _to_int(cfg.get('Umbral_Alerta_pct', -5)) if cfg else -5
        df_anom_med = df_med2[df_med2['Desv_CB_pct'] <= ua].copy()

    total    = len(df_fallas)
    cortes   = len(df_fallas[df_fallas['Amperios'] == 0])
    alertas  = len(df_fallas[df_fallas['Amperios'] > 0])
    reinc    = len(df_fallas.groupby(['Inversor','Caja','String']).filter(lambda x: len(x) > 1))

    # ── Portada ──
    pdf.set_fill_color(192, 57, 43)
    pdf.rect(0, 28, 210, 28, 'F')
    pdf.set_font("Arial","B",16); pdf.set_text_color(255,255,255)
    pdf.set_xy(10, 33)
    pdf.cell(0, 10, clean_text(f"INFORME DE FALLAS — {planta_nombre}"), 0, 1, 'C')
    pdf.set_font("Arial","",10); pdf.set_text_color(220,220,220)
    pdf.set_x(10)
    pdf.cell(0, 8, "Elaborado por Mundo Solar SpA", 0, 1, 'C')
    pdf.ln(12)

    # ── KPIs ──
    prom_lbl = f"{prom_med:.3f}" if prom_med is not None else "-"
    kpis = [
        ("Fusibles registrados", str(total),    "1A3A5C"),
        ("I Prom. Planta (Med.)", prom_lbl,     "2E6DA4"),
        ("OC (0A)",           str(cortes),  "C0392B"),
        ("Strings c/Desv.",       str(len(df_anom_med)), "E67E22"),
    ]
    col_w = 45; x0 = 10
    for lbl, val, color in kpis:
        r,g,b = int(color[:2],16), int(color[2:4],16), int(color[4:],16)
        pdf.set_fill_color(r,g,b); pdf.set_text_color(255,255,255)
        pdf.set_xy(x0, pdf.get_y())
        pdf.set_font("Arial","B",18); pdf.cell(col_w-3, 12, val, 0, 0, 'C', True)
        pdf.set_xy(x0, pdf.get_y()+12)
        pdf.set_font("Arial","",8)
        r2 = min(r+30,255); g2 = min(g+30,255); b2 = min(b+30,255)
        pdf.set_fill_color(r2,g2,b2)
        pdf.cell(col_w-3, 7, clean_text(lbl), 0, 0, 'C', True)
        x0 += col_w
    pdf.ln(22); pdf.set_text_color(0,0,0); pdf.ln(4)

    # ── Tabla fusibles ──
    pdf.set_font("Arial","B",9)
    pdf.set_fill_color(192,57,43); pdf.set_text_color(255,255,255)
    pdf.cell(0, 7, clean_text(f"Fusibles registrados — {total} eventos"), 0, 1, 'L', True)
    pdf.ln(1)
    pdf.set_fill_color(46,109,164)
    hdrs = ["Fecha","Inv","Caja","String","Pol","A","Irr.","Tipo","Nota"]
    ws_  = [20, 14, 12, 12, 22, 11, 14, 30, 45]
    for h, w in zip(hdrs, ws_): pdf.cell(w, 7, h, 1, 0, 'C', True)
    pdf.ln()
    pdf.set_font("Arial","",7.5); pdf.set_text_color(0,0,0)
    color_tipo = {
        'OC (0A)': (250,219,216), 'Fallo grave (<-30%)': (235,180,180),
        'Critico (-15% a -30%)': (250,219,216), 'Alerta (-5% a -15%)': (254,249,231),
        'Alerta (4-6A)': (254,249,231), 'Critico (<4A)': (250,219,216),
    }
    for i, (_, r) in enumerate(df_fallas.iterrows()):
        tipo = str(r.get('Tipo',''))
        bg = color_tipo.get(tipo, (247,249,252) if i%2==0 else (255,255,255))
        pdf.set_fill_color(*bg)
        fecha_str = r['Fecha'].strftime('%d/%m/%Y') if pd.notna(r.get('Fecha')) else ''
        irr_val = _to_float(r.get('Irradiancia_Wm2',0))
        irr_str = str(int(irr_val)) if irr_val > 0 else '-'
        vals = [fecha_str, str(r.get('Inversor','')), str(r.get('Caja','')),
                str(r.get('String','')), str(r.get('Polaridad',''))[:3],
                f"{_to_float(r.get('Amperios',0)):.1f}", irr_str,
                clean_text(tipo)[:18], clean_text(str(r.get('Nota','')))[:25]]
        for v, w in zip(vals, ws_): pdf.cell(w, 6.5, clean_text(str(v)), 1, 0, 'L', True)
        pdf.ln()
        if pdf.get_y() > 262:
            pdf.add_page()
            pdf.set_font("Arial","B",7.5); pdf.set_fill_color(46,109,164); pdf.set_text_color(255,255,255)
            for h, w in zip(hdrs, ws_): pdf.cell(w,7,h,1,0,'C',True)
            pdf.ln(); pdf.set_font("Arial","",7.5); pdf.set_text_color(0,0,0)

    # ── Tabla strings con desviación (desde mediciones) ──
    if not df_anom_med.empty:
        pdf.ln(5)
        pdf.set_font("Arial","B",9)
        pdf.set_fill_color(231,118,26); pdf.set_text_color(255,255,255)
        pdf.cell(0, 7, clean_text(f"Strings con desviacion respecto a su CB — {len(df_anom_med)} detectados"), 0, 1, 'L', True)
        pdf.ln(1)
        pdf.set_fill_color(46,109,164)
        hdrs2 = ["Equipo (CB)","String","I medida (A)","Prom. CB (A)","Desv. %","Estado"]
        ws2   = [40, 22, 28, 28, 22, 40]
        for h, w in zip(hdrs2, ws2): pdf.cell(w, 7, h, 1, 0, 'C', True)
        pdf.ln()
        pdf.set_font("Arial","",7.5); pdf.set_text_color(0,0,0)
        sid_col = 'String ID' if 'String ID' in df_anom_med.columns else 'String_ID'
        for i, (_, r) in enumerate(df_anom_med.head(50).iterrows()):
            desv = _to_float(r.get('Desv_CB_pct', 0))
            diag = str(r.get('Diagnostico',''))
            if 'CRITICO' in diag or 'CORTE' in diag: pdf.set_fill_color(250,219,216)
            elif 'ALERTA' in diag:                    pdf.set_fill_color(254,249,231)
            else: pdf.set_fill_color(247,249,252) if i%2==0 else pdf.set_fill_color(255,255,255)
            vals2 = [str(r.get('Equipo','')), str(r.get(sid_col,'')),
                     f"{_to_float(r.get('Amperios',0)):.2f}",
                     f"{_to_float(r.get('Promedio_Caja',0)):.3f}",
                     f"{desv:+.1f}%", clean_text(diag)[:20]]
            for v, w in zip(vals2, ws2): pdf.cell(w, 6.5, clean_text(v), 1, 0, 'C', True)
            pdf.ln()
            if pdf.get_y() > 262:
                pdf.add_page()
                pdf.set_font("Arial","B",7.5); pdf.set_fill_color(46,109,164); pdf.set_text_color(255,255,255)
                for h, w in zip(hdrs2, ws2): pdf.cell(w,7,h,1,0,'C',True)
                pdf.ln(); pdf.set_font("Arial","",7.5); pdf.set_text_color(0,0,0)

    # ── Sección Recurrencia en PDF ──
    conteo_pdf, recur_pdf, cb_pdf = _calcular_recurrencia_df(df_fallas)
    if not conteo_pdf.empty:
        pdf.ln(5)
        pdf.set_font("Arial","B",9)
        pdf.set_fill_color(192,57,43); pdf.set_text_color(255,255,255)
        n_rec_pdf = len(recur_pdf)
        n_ub_pdf  = len(conteo_pdf)
        tasa_pdf  = round(n_rec_pdf / n_ub_pdf * 100, 1) if n_ub_pdf > 0 else 0
        pdf.cell(0, 7, clean_text(f"Recurrencia de Fallos — {n_rec_pdf} strings con mas de 1 falla ({tasa_pdf}%)"), 0, 1, 'L', True)
        pdf.ln(1)

        # Mini KPIs
        pdf.set_font("Arial","B",8); pdf.set_text_color(255,255,255)
        kpis_p = [("Strings afect.", str(n_ub_pdf), "1A3A5C"),
                  ("Con recurrencia", str(n_rec_pdf), "C0392B"),
                  ("Tasa recurrencia", f"{tasa_pdf}%", "E67E22"),
                  ("Max fallas/str", str(int(conteo_pdf['N_Fallas'].max())), "8B0000")]
        for lbl_p, val_p, col_p in kpis_p:
            r_p,g_p,b_p = int(col_p[:2],16),int(col_p[2:4],16),int(col_p[4:],16)
            pdf.set_fill_color(r_p,g_p,b_p)
            pdf.cell(42, 10, val_p, 0, 0, 'C', True)
        pdf.ln(10)
        pdf.set_font("Arial","",7); pdf.set_text_color(0,0,0)
        for lbl_p, val_p, col_p in kpis_p:
            pdf.cell(42, 5, clean_text(lbl_p), 0, 0, 'C')
        pdf.ln(7)

        if not recur_pdf.empty:
            pdf.set_font("Arial","B",8)
            pdf.set_fill_color(46,109,164); pdf.set_text_color(255,255,255)
            hdrs_r = ["Ubicacion","N Fallas","Categoria","Primera","Ultima","MTBF(d)"]
            ws_r   = [55, 16, 28, 22, 22, 17]
            for h,w in zip(hdrs_r, ws_r): pdf.cell(w,7,h,1,0,'C',True)
            pdf.ln()
            pdf.set_font("Arial","",7.5); pdf.set_text_color(0,0,0)
            color_r = {'Sin recurrencia':(247,249,252),'Recurrente (2x)':(254,249,231),
                       'Critico (3-4x)':(250,219,216),'Cronico (5+)':(245,183,177)}
            for _, row in recur_pdf.head(30).iterrows():
                cat = str(row.get('Categoria',''))
                bg  = color_r.get(cat,(247,249,252))
                pdf.set_fill_color(*bg)
                prim = row['Primera'].strftime('%d/%m/%y') if pd.notna(row.get('Primera')) else '-'
                ult  = row['Ultima'].strftime('%d/%m/%y')  if pd.notna(row.get('Ultima'))  else '-'
                mtbf = f"{row['MTBF_dias']:.1f}" if row.get('MTBF_dias') is not None else '-'
                vals_r = [str(row['Ubicacion'])[:28], str(int(row['N_Fallas'])),
                          clean_text(cat)[:18], prim, ult, mtbf]
                for v,w in zip(vals_r, ws_r): pdf.cell(w,6.5,clean_text(v),1,0,'C',True)
                pdf.ln()
                if pdf.get_y() > 265:
                    pdf.add_page()
                    pdf.set_font("Arial","B",7.5); pdf.set_fill_color(46,109,164); pdf.set_text_color(255,255,255)
                    for h,w in zip(hdrs_r, ws_r): pdf.cell(w,7,h,1,0,'C',True)
                    pdf.ln(); pdf.set_font("Arial","",7.5); pdf.set_text_color(0,0,0)

    _out = pdf.output(dest='S')
    if isinstance(_out, bytes): return _out
    if isinstance(_out, bytearray): return bytes(_out)
    return _out.encode('latin-1')


def _narrativa_auto(df_proc, planta_nombre):
    """Genera narrativa automática de diagnóstico."""
    total   = len(df_proc)
    normales= len(df_proc[df_proc['Diagnostico']=='NORMAL'])
    alertas = len(df_proc[df_proc['Diagnostico']=='ALERTA'])
    criticos= len(df_proc[df_proc['Diagnostico'].isin(['CRÍTICO','OC (0A)'])])
    cortes  = len(df_proc[df_proc['Diagnostico']=='OC (0A)'])
    salud   = (normales/total*100) if total>0 else 0
    prom_g  = df_proc['Amperios'].mean()

    df_cb   = df_proc.groupby('Equipo')['Amperios'].mean()
    mejor   = df_cb.idxmax(); mejor_v = df_cb.max()
    peor    = df_cb.idxmin(); peor_v  = df_cb.min()

    estado_txt = ("optimas, sin anomalias criticas detectadas" if criticos==0 and alertas==0
                  else f"aceptables con {alertas} strings en alerta" if criticos==0
                  else f"requieren atencion con {criticos} strings criticos y {alertas} en alerta")

    texto = (
        f"Durante la inspeccion tecnica realizada en la planta {planta_nombre}, "
        f"se evaluaron {total} strings con una corriente promedio global de {prom_g:.2f} A. "
        f"Las condiciones operativas se presentan {estado_txt}. "
    )
    if cortes > 0:
        texto += f"Se identificaron {cortes} strings sin generacion (circuito abierto OC (0A)), los cuales requieren inspeccion urgente. "
    if alertas > 0:
        texto += f"Adicionalmente, {alertas} strings presentan desviacion moderada respecto al promedio de su caja, sugiriendo suciedad, sombra parcial o degradacion inicial. "
    texto += (
        f"El indice de salud general de la planta es {salud:.1f}%. "
        f"La caja de mejor desempeno es {mejor} con {mejor_v:.2f} A promedio, "
        f"mientras que {peor} registro el valor mas bajo con {peor_v:.2f} A, "
        f"lo que sugiere revision focalizada en ese sector."
    )
    return texto

def generar_pdf_mediciones(planta_nombre, df, cfg=None, restriccion_mw=None, capacidad_mw=0, num_inversores=1, df_fallas=None):
    """PDF profesional completo: portada, KPIs, narrativa, graficos, tabla anomalias, acciones."""
    import tempfile, os
    pdf = PDF()

    isc_nom     = _to_float(cfg.get('Isc_STC_A', 9.07)) if cfg else 9.07
    impp_nom    = _to_float(cfg.get('Impp_STC_A', 8.68)) if cfg else 8.68
    pmax        = _to_float(cfg.get('Pmax_W', 320)) if cfg else 320
    panels      = _to_int(cfg.get('Panels_por_String', 30)) if cfg else 30
    ua          = _to_int(cfg.get('Umbral_Alerta_pct', -5)) if cfg else -5
    uc          = _to_int(cfg.get('Umbral_Critico_pct', -10)) if cfg else -10
    capacidad   = str(cfg.get('Capacidad', '')) if cfg else ''
    modulo      = str(cfg.get('Modulo', '')) if cfg else ''
    irradiancia = 698
    isc_ref     = round(isc_nom * irradiancia / 1000, 3)

    # Restricción CEN
    rest_activa = restriccion_mw and capacidad_mw and restriccion_mw > 0 and capacidad_mw > 0
    factor_rest = (restriccion_mw / capacidad_mw) if rest_activa else 1.0
    mw_inv_rest = (restriccion_mw / num_inversores) if (rest_activa and num_inversores > 0) else restriccion_mw

    df_proc = analizar_mediciones(df, ua=ua, uc=uc,
        restriccion_mw=restriccion_mw if rest_activa else None,
        capacidad_mw=capacidad_mw if rest_activa else None)
    if 'String ID' not in df_proc.columns and 'String_ID' in df_proc.columns:
        df_proc.rename(columns={'String_ID':'String ID'}, inplace=True)

    total    = len(df_proc)
    normales = len(df_proc[df_proc['Diagnostico']=='NORMAL'])
    alertas  = len(df_proc[df_proc['Diagnostico']=='ALERTA'])
    criticos = len(df_proc[df_proc['Diagnostico'].isin(['CRÍTICO','OC (0A)'])])
    salud    = (normales/total*100) if total>0 else 0
    prom_g   = df_proc['Amperios'].mean()
    df_anom  = df_proc[df_proc['Diagnostico']!='NORMAL'].sort_values('Desv_CB_pct', ascending=True)
    cb_sum   = df_proc.groupby('Equipo')['Amperios'].agg(['mean','min','max','std']).reset_index()
    cb_sum.columns = ['Equipo','I_media','I_min','I_max','Istd']

    # Texto de restricción CEN para incluir en PDF
    nota_restriccion = None
    if rest_activa:
        nota_restriccion = (
            f"NOTA: Esta campaña fue realizada bajo instruccion operacional del CEN "
            f"con limitacion de inyeccion. La planta opero al {factor_rest*100:.1f}% "
            f"de su capacidad ({restriccion_mw:.1f} MW de {capacidad_mw:.1f} MW totales), "
            f"distribuyendo {mw_inv_rest:.2f} MW por inversor ({num_inversores} inversores). "
            f"El Isc_ref esperado fue ajustado proporcionalmente. "
            f"Los diagnosticos reflejan el desempeno real bajo condicion restringida."
        )

    # ════════════════════════════
    # PÁGINA 1 — PORTADA
    # ════════════════════════════
    pdf.add_page()

    # Banner azul superior
    pdf.set_fill_color(26, 58, 92)
    pdf.rect(0, 15, 210, 40, 'F')
    pdf.set_fill_color(244, 196, 48)
    pdf.rect(0, 52, 210, 3, 'F')

    pdf.set_font("Arial","B",20); pdf.set_text_color(255,255,255)
    pdf.set_xy(10, 20)
    pdf.cell(0, 12, "INFORME TECNICO DE MEDICIONES DE STRINGS", 0, 1, 'C')
    pdf.set_font("Arial","",11); pdf.set_text_color(168,209,245)
    pdf.set_x(10)
    pdf.cell(0, 8, clean_text(f"Mundo Solar SpA  |  pMGD Solar — {planta_nombre}  |  {capacidad}"), 0, 1, 'C')

    pdf.set_text_color(0,0,0); pdf.ln(10)

    # Dos columnas: datos proyecto | parametros STC
    pdf.set_font("Arial","B",10); pdf.set_fill_color(26,58,92); pdf.set_text_color(255,255,255)
    pdf.cell(93, 8, "DATOS DEL PROYECTO", 0, 0, 'C', True)
    pdf.cell(4, 8, "", 0, 0)
    pdf.cell(93, 8, "PARAMETROS STC", 0, 1, 'C', True)
    pdf.set_text_color(0,0,0)

    datos_proy = [("Planta:", clean_text(f"pMGD {planta_nombre}")),
                  ("Capacidad:", clean_text(capacidad)),
                  ("Modulo FV:", clean_text(f"{modulo} / {int(pmax)} Wp")),
                  ("Paneles/string:", str(panels)),
                  ("Total strings:", str(total)),
                  ("Irradiancia:", f"~{irradiancia} W/m2")]
    datos_stc  = [("Isc nominal:", f"{isc_nom} A"),
                  ("Impp nominal:", f"{impp_nom} A"),
                  ("Isc corregida:", f"{isc_ref} A"),
                  ("Umbral ALERTA:", f"< {ua}%"),
                  ("Umbral CRITICO:", f"< {uc}%"),
                  ("", "")]

    for (lp, vp), (ls, vs) in zip(datos_proy, datos_stc):
        pdf.set_font("Arial","B",9); pdf.set_fill_color(216,232,245)
        pdf.cell(38, 7, clean_text(lp), 0, 0, 'L', True)
        pdf.set_font("Arial","",9); pdf.set_fill_color(247,249,252)
        pdf.cell(55, 7, clean_text(vp), 0, 0, 'L', True)
        pdf.cell(4, 7, "", 0, 0)
        pdf.set_font("Arial","B",9); pdf.set_fill_color(216,232,245)
        pdf.cell(38, 7, clean_text(ls), 0, 0, 'L', True)
        pdf.set_font("Arial","",9); pdf.set_fill_color(247,249,252)
        pdf.cell(55, 7, clean_text(vs), 0, 1, 'L', True)
    pdf.ln(5)

    # KPIs grandes con color
    pdf.set_font("Arial","B",10); pdf.set_fill_color(26,58,92); pdf.set_text_color(255,255,255)
    pdf.cell(0, 8, "RESUMEN EJECUTIVO", 0, 1, 'C', True)
    pdf.set_text_color(0,0,0)

    kpi_data = [
        ("Total Strings", str(total),          "D8E8F5", "1F5C8B"),
        ("NORMAL",         str(normales),        "D5F5E3", "1E8449"),
        ("ALERTA",         str(alertas),         "FEF9E7", "D4AC0D"),
        ("CRITICO/CORTE",  str(criticos),        "FADBD8", "C0392B"),
        ("I Media Global", f"{prom_g:.3f} A",   "D8E8F5", "1F5C8B"),
        ("Salud Planta",   f"{salud:.1f}%",      "D8E8F5", "1A3A5C"),
    ]
    cb_min = cb_sum.loc[cb_sum['I_media'].idxmin()] if not cb_sum.empty else None
    if cb_min is not None:
        kpi_data.append(("CB mas baja", clean_text(f"{cb_min['Equipo']} ({cb_min['I_media']:.3f}A)"), "FADBD8", "C0392B"))
    if nota_restriccion:
        kpi_data.insert(1, ("Restriccion CEN",
            clean_text(f"{restriccion_mw:.1f} MW / {capacidad_mw:.1f} MW ({factor_rest*100:.1f}%) · {mw_inv_rest:.2f} MW/inv"),
            "FFF3CD", "856404"))

    for lbl, val, bg, fg in kpi_data:
        r_bg = (int(bg[:2],16), int(bg[2:4],16), int(bg[4:],16))
        r_fg = (int(fg[:2],16), int(fg[2:4],16), int(fg[4:],16))
        pdf.set_font("Arial","B",9); pdf.set_fill_color(*r_bg); pdf.set_text_color(70,70,70)
        pdf.cell(65, 7, clean_text(lbl), 0, 0, 'L', True)
        pdf.set_font("Arial","B",10); pdf.set_text_color(*r_fg)
        pdf.cell(125, 7, clean_text(val), 0, 1, 'C', True)
    pdf.set_text_color(0,0,0)

    # ════════════════════════════
    # PÁGINA 2 — NARRATIVA + GRÁFICO BARRAS
    # ════════════════════════════
    pdf.add_page()
    pdf.set_font("Arial","B",12); pdf.set_fill_color(26,58,92); pdf.set_text_color(255,255,255)
    pdf.cell(0, 9, "1. ANALISIS EJECUTIVO DE PERFORMANCE", 0, 1, 'C', True)
    pdf.set_text_color(0,0,0); pdf.ln(3)

    # Nota restricción CEN (si aplica)
    if nota_restriccion:
        pdf.set_fill_color(255, 243, 205); pdf.set_draw_color(230, 160, 0)
        pdf.set_font("Arial","B",9); pdf.set_text_color(120, 80, 0)
        pdf.cell(0, 7, "AVISO: CAMPANA REALIZADA BAJO RESTRICCION OPERACIONAL CEN", 1, 1, 'C', True)
        pdf.set_font("Arial","",9); pdf.set_text_color(80, 60, 0)
        pdf.set_fill_color(255, 250, 220)
        pdf.multi_cell(0, 5, clean_text(nota_restriccion), 1, 'L', True)
        pdf.set_text_color(0,0,0); pdf.set_fill_color(255,255,255)
        pdf.ln(3)

    narrativa = _narrativa_auto(df_proc, planta_nombre)
    pdf.set_font("Arial","",10)
    pdf.multi_cell(0, 6, clean_text(narrativa))
    pdf.ln(5)

    # Semaforo visual
    pdf.set_font("Arial","B",10); pdf.set_fill_color(26,58,92); pdf.set_text_color(255,255,255)
    pdf.cell(0, 8, "SEMAFORO DE ESTADO", 0, 1, 'C', True)
    pdf.set_text_color(0,0,0)
    for lbl, n, bg in [("NORMAL", normales, (30,132,73)),
                        ("ALERTA", alertas,  (212,172,13)),
                        ("CRITICO/CORTE", criticos, (192,57,43))]:
        pct = (n/total*100) if total>0 else 0
        pdf.set_font("Arial","B",9); pdf.set_fill_color(*bg); pdf.set_text_color(255,255,255)
        pdf.cell(40, 8, clean_text(lbl), 0, 0, 'C', True)
        pdf.set_fill_color(240,240,240); pdf.set_text_color(0,0,0)
        pdf.set_font("Arial","",9)
        pdf.cell(100, 8, f"{n} strings  ({pct:.1f}%)", 0, 0, 'L', True)
        bar_w = int(pct * 0.5)
        pdf.set_fill_color(*bg)
        pdf.cell(bar_w if bar_w>0 else 1, 8, "", 0, 0, 'L', True)
        pdf.set_fill_color(220,220,220)
        pdf.cell(50-bar_w if 50-bar_w>0 else 1, 8, "", 0, 1, 'L', True)
    pdf.ln(5)

    # Gráfico de barras por CB (Plotly → PNG embebido)
    pdf.set_font("Arial","B",12); pdf.set_fill_color(26,58,92); pdf.set_text_color(255,255,255)
    pdf.cell(0, 9, "2. CORRIENTE MEDIA POR COMBINER BOX", 0, 1, 'C', True)
    pdf.set_text_color(0,0,0); pdf.ln(2)
    try:
        import plotly.graph_objects as go_pdf
        colors_bar = []
        for _, row in cb_sum.iterrows():
            desv = ((row['I_media'] - prom_g) / prom_g * 100) if prom_g > 0 else 0
            # Colores semaforo exactos para las barras de cada Combiner Box
            colors_bar.append('#C0392B' if desv <= uc else '#E67E22' if desv <= ua else '#1E8449')
            
        fig_bar = go_pdf.Figure(go_pdf.Bar(
            x=cb_sum['Equipo'].str.replace('Inv-1>','',regex=False),
            y=cb_sum['I_media'].round(3),
            marker=dict(color=colors_bar),
            text=cb_sum['I_media'].round(3), textposition='outside',
        ))
        fig_bar.add_hline(y=prom_g, line_dash='dash', line_color='#1A3A5C',
                          annotation_text=f'Media: {prom_g:.3f}A')
        fig_bar.update_layout(
            height=320, width=750, showlegend=False,
            plot_bgcolor='white', paper_bgcolor='white',
            margin=dict(t=30,b=50,l=50,r=30),
            xaxis=dict(tickangle=-30, showgrid=False),
            yaxis=dict(showgrid=True, gridcolor='#f0f0f0'),
            font=dict(family='Arial', size=11)
        )
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tf:
            fig_bar.write_image(tf.name, scale=2)
            pdf.image(tf.name, x=10, w=185)
        os.unlink(tf.name)
    except Exception as e:
        pdf.set_font("Arial","I",9); pdf.set_text_color(150,150,150)
        pdf.cell(0, 8, f"[Grafico no disponible: {e}]", 0, 1, 'C')
        pdf.set_text_color(0,0,0)
    pdf.ln(3)

    # ════════════════════════════
    # PÁGINA 3 — BOXPLOT + RESUMEN CB
    # ════════════════════════════
    # Boxplot
    try:
        import plotly.express as px_pdf
        
        # 1. Creamos una paleta de colores vibrantes
        paleta = ['#e6194b', '#3cb44b', '#ffe119', '#4363d8', '#f58231', 
                   '#911eb4', '#46f0f0', '#f032e6', '#bcf60c', '#fabebe', 
                   '#008080', '#e6beff', '#9a6324', '#fffac8', '#800000', 
                   '#aaffc3', '#808000', '#ffd8b1', '#000075', '#808080'] * 5
        
        equipos_unicos = df_proc['Equipo'].unique()
        mapa_colores = {eq: paleta[i] for i, eq in enumerate(equipos_unicos)}

        # 2. Generamos el gráfico asignando explícitamente el mapa
        fig_box = px_pdf.box(df_proc, x='Equipo', y='Amperios', color='Equipo', 
                             color_discrete_map=mapa_colores, points='outliers')
        
        # 3. FORZAMOS a que cada caja se pinte por dentro y por fuera (Anti-Kaleido bug)
        for trazo in fig_box.data:
            color_asignado = mapa_colores.get(trazo.name, '#1F5C8B')
            trazo.fillcolor = color_asignado
            trazo.line.color = color_asignado
            trazo.marker.color = color_asignado

        fig_box.update_layout(
            height=300, width=750, showlegend=False,
            plot_bgcolor='white', paper_bgcolor='white',
            margin=dict(t=20,b=50,l=50,r=20),
            xaxis=dict(tickangle=-30, showgrid=False, title=''),
            yaxis=dict(showgrid=True, gridcolor='#f0f0f0', title='Corriente (A)'),
            font=dict(family='Arial', size=10)
        )
        fig_box.update_xaxes(ticktext=[e.replace('Inv-1>','') for e in equipos_unicos],
                             tickvals=equipos_unicos)
                             
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tf:
            fig_box.write_image(tf.name, scale=2)
            pdf.image(tf.name, x=10, w=185)
        import os
        os.unlink(tf.name)
    except Exception as e:
        pdf.set_font("Arial","I",9); pdf.set_text_color(150,150,150)
        pdf.cell(0, 8, f"[Boxplot no disponible: {e}]", 0, 1, 'C')
        pdf.set_text_color(0,0,0)

    # Tabla resumen CB
    pdf.set_font("Arial","B",9); pdf.set_fill_color(46,109,164); pdf.set_text_color(255,255,255)
    h_cols = ["Caja","I media","I min","I max","Std Dev","Desv.Global%","Alertas","Criticos"]
    w_cols = [38,20,18,18,18,25,18,18]
    for h,w in zip(h_cols,w_cols): pdf.cell(w,8,h,1,0,'C',True)
    pdf.ln(); pdf.set_text_color(0,0,0)

    for i,(_, r) in enumerate(cb_sum.iterrows()):
        desv_g = ((r['I_media']-prom_g)/prom_g*100) if prom_g>0 else 0
        n_al = len(df_proc[(df_proc['Equipo']==r['Equipo']) & (df_proc['Diagnostico']=='ALERTA')])
        n_cr = len(df_proc[(df_proc['Equipo']==r['Equipo']) & (df_proc['Diagnostico'].isin(['CRÍTICO','OC (0A)']))])
        bg_alt = (247,249,252) if i%2==0 else (255,255,255)
        bg_row = (250,219,216) if n_cr>0 else (254,249,231) if n_al>0 else bg_alt
        pdf.set_fill_color(*bg_row)
        pdf.set_font("Arial","",8)
        vals = [str(r['Equipo']).replace('Inv-1>',''), f"{r['I_media']:.3f}", f"{r['I_min']:.2f}",
                f"{r['I_max']:.2f}", f"{r['Istd']:.3f}", f"{desv_g:+.2f}%", str(n_al), str(n_cr)]
        for v,w in zip(vals,w_cols): pdf.cell(w,7,clean_text(v),1,0,'C',True)
        pdf.ln()

    # ════════════════════════════
    # PÁGINA 4 — STRINGS FUERA DE RANGO + ACCIONES
    # ════════════════════════════
    pdf.add_page()
    pdf.set_font("Arial","B",12); pdf.set_fill_color(192,57,43); pdf.set_text_color(255,255,255)
    pdf.cell(0, 9, clean_text(f"4. STRINGS FUERA DE RANGO — {len(df_anom)} identificados"), 0, 1, 'C', True)
    pdf.set_text_color(0,0,0); pdf.ln(3)

    if df_anom.empty:
        pdf.set_font("Arial","",10); pdf.set_fill_color(213,245,227); pdf.set_text_color(21,84,46)
        pdf.cell(0,10,"No se detectaron desviaciones criticas en esta inspeccion.",1,1,'C',True)
        pdf.set_text_color(0,0,0)
    else:
        def causa_pdf(d):
            if d <= uc: return "Modulo defectuoso / conector MC4 danado o bypass activado"
            if d <= -7: return "Modulo degradado / suciedad intensa o sombra parcial"
            return "Suciedad leve / sombra o degradacion inicial"
        def accion_pdf(d):
            if d <= uc: return "Inspeccion urgente + termografia + curva I-V"
            if d <= -7: return "Inspeccion + limpieza + revision conectores"
            return "Monitorear + limpieza preventiva"

        pdf.set_font("Arial","B",8); pdf.set_fill_color(46,109,164); pdf.set_text_color(255,255,255)
        h2 = ["#","Caja","String","I (A)","Prom.CB","Desv%","Estado","Causa","Accion"]
        w2 = [8, 30, 16, 14, 14, 14, 20, 42, 42]
        for h,w in zip(h2,w2): pdf.cell(w,8,h,1,0,'C',True)
        pdf.ln(); pdf.set_text_color(0,0,0)

        for i,(_, r) in enumerate(df_anom.iterrows(), 1):
            diag = str(r.get('Diagnostico',''))
            desv = _to_float(r.get('Desv_CB_pct',0))
            if 'CRITICO' in diag or 'CORTE' in diag: pdf.set_fill_color(250,219,216)
            else:                                      pdf.set_fill_color(254,249,231)
            pdf.set_font("Arial","",7)
            vals = [str(i), str(r.get('Equipo','')).replace('Inv-1>',''),
                    str(r.get('String ID','')),
                    f"{_to_float(r.get('Amperios',0)):.2f}",
                    f"{_to_float(r.get('Promedio_Caja',0)):.3f}",
                    f"{desv:+.1f}%", clean_text(diag),
                    clean_text(causa_pdf(desv)), clean_text(accion_pdf(desv))]
            for v,w in zip(vals,w2): pdf.cell(w,7,v[:35],1,0,'C',True)
            pdf.ln()
            if pdf.get_y() > 265:
                pdf.add_page()
                pdf.set_font("Arial","B",8); pdf.set_fill_color(46,109,164); pdf.set_text_color(255,255,255)
                for h,w in zip(h2,w2): pdf.cell(w,8,h,1,0,'C',True)
                pdf.ln(); pdf.set_text_color(0,0,0)

    # ── Sección Recurrencia en PDF ──
    conteo_pdf, recur_pdf, cb_pdf = _calcular_recurrencia_df(df_fallas)
    if not conteo_pdf.empty:
        pdf.ln(5)
        pdf.set_font("Arial","B",9)
        pdf.set_fill_color(192,57,43); pdf.set_text_color(255,255,255)
        n_rec_pdf = len(recur_pdf)
        n_ub_pdf  = len(conteo_pdf)
        tasa_pdf  = round(n_rec_pdf / n_ub_pdf * 100, 1) if n_ub_pdf > 0 else 0
        pdf.cell(0, 7, clean_text(f"Recurrencia de Fallos — {n_rec_pdf} strings con mas de 1 falla ({tasa_pdf}%)"), 0, 1, 'L', True)
        pdf.ln(1)

        # Mini KPIs
        pdf.set_font("Arial","B",8); pdf.set_text_color(255,255,255)
        kpis_p = [("Strings afect.", str(n_ub_pdf), "1A3A5C"),
                  ("Con recurrencia", str(n_rec_pdf), "C0392B"),
                  ("Tasa recurrencia", f"{tasa_pdf}%", "E67E22"),
                  ("Max fallas/str", str(int(conteo_pdf['N_Fallas'].max())), "8B0000")]
        for lbl_p, val_p, col_p in kpis_p:
            r_p,g_p,b_p = int(col_p[:2],16),int(col_p[2:4],16),int(col_p[4:],16)
            pdf.set_fill_color(r_p,g_p,b_p)
            pdf.cell(42, 10, val_p, 0, 0, 'C', True)
        pdf.ln(10)
        pdf.set_font("Arial","",7); pdf.set_text_color(0,0,0)
        for lbl_p, val_p, col_p in kpis_p:
            pdf.cell(42, 5, clean_text(lbl_p), 0, 0, 'C')
        pdf.ln(7)

        if not recur_pdf.empty:
            pdf.set_font("Arial","B",8)
            pdf.set_fill_color(46,109,164); pdf.set_text_color(255,255,255)
            hdrs_r = ["Ubicacion","N Fallas","Categoria","Primera","Ultima","MTBF(d)"]
            ws_r   = [55, 16, 28, 22, 22, 17]
            for h,w in zip(hdrs_r, ws_r): pdf.cell(w,7,h,1,0,'C',True)
            pdf.ln()
            pdf.set_font("Arial","",7.5); pdf.set_text_color(0,0,0)
            color_r = {'Sin recurrencia':(247,249,252),'Recurrente (2x)':(254,249,231),
                       'Critico (3-4x)':(250,219,216),'Cronico (5+)':(245,183,177)}
            for _, row in recur_pdf.head(30).iterrows():
                cat = str(row.get('Categoria',''))
                bg  = color_r.get(cat,(247,249,252))
                pdf.set_fill_color(*bg)
                prim = row['Primera'].strftime('%d/%m/%y') if pd.notna(row.get('Primera')) else '-'
                ult  = row['Ultima'].strftime('%d/%m/%y')  if pd.notna(row.get('Ultima'))  else '-'
                mtbf = f"{row['MTBF_dias']:.1f}" if row.get('MTBF_dias') is not None else '-'
                vals_r = [str(row['Ubicacion'])[:28], str(int(row['N_Fallas'])),
                          clean_text(cat)[:18], prim, ult, mtbf]
                for v,w in zip(vals_r, ws_r): pdf.cell(w,6.5,clean_text(v),1,0,'C',True)
                pdf.ln()
                if pdf.get_y() > 265:
                    pdf.add_page()
                    pdf.set_font("Arial","B",7.5); pdf.set_fill_color(46,109,164); pdf.set_text_color(255,255,255)
                    for h,w in zip(hdrs_r, ws_r): pdf.cell(w,7,h,1,0,'C',True)
                    pdf.ln(); pdf.set_font("Arial","",7.5); pdf.set_text_color(0,0,0)

    _out = pdf.output(dest='S')
    if isinstance(_out, bytes): return _out
    if isinstance(_out, bytearray): return bytes(_out)
    return _out.encode('latin-1')

# ══════════════════════════════════════════════════════════════
# EXCEL ENGINE — Replica exacta del informe de referencia + hoja Fallas
# ══════════════════════════════════════════════════════════════
def _fill(c): return PatternFill('solid',start_color=c,fgColor=c)
def _fnt(bold=False,size=10,color='000000',italic=False):
    return Font(name='Arial',bold=bold,size=size,color=color,italic=italic)
def _aln(h='center',v='center',wrap=False):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def _brd():
    t=Side(style='thin',color='CCCCCC')
    return Border(left=t,right=t,top=t,bottom=t)
def _hdr(cell,text,bg='1A3A5C',fg='FFFFFF',size=10):
    cell.value=text; cell.font=_fnt(bold=True,size=size,color=fg)
    cell.fill=_fill(bg); cell.alignment=_aln(wrap=True); cell.border=_brd()
def _dc(cell,val,bg='FFFFFF',bold=False,h='center',fmt=None,color='000000'):
    cell.value=val; cell.font=_fnt(bold=bold,size=10,color=color)
    cell.fill=_fill(bg); cell.alignment=_aln(h=h); cell.border=_brd()
    if fmt: cell.number_format=fmt


def _calcular_recurrencia_df(df_fallas):
    """Calcula recurrencia de fallos. Retorna (conteo_df, recurrentes_df, cb_rank_df)."""
    if df_fallas is None or df_fallas.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    df = df_fallas.copy()
    df['_ubic'] = (df['Inversor'].astype(str) + '>CB-' +
                   df['Caja'].astype(str) + '>Str-' +
                   df['String'].astype(str))
    conteo = (df.groupby('_ubic')
                .agg(N_Fallas=('ID','count'),
                     Primera=('Fecha','min'),
                     Ultima=('Fecha','max'),
                     Inversor=('Inversor','first'),
                     Caja=('Caja','first'),
                     String=('String','first'))
                .reset_index()
                .rename(columns={'_ubic':'Ubicacion'})
                .sort_values('N_Fallas', ascending=False))
    conteo['Dias'] = (conteo['Ultima'] - conteo['Primera']).dt.days.fillna(0).astype(int)
    def _cat(n):
        if n == 1:  return 'Sin recurrencia'
        if n == 2:  return 'Recurrente (2x)'
        if n <= 4:  return 'Critico (3-4x)'
        return 'Cronico (5+)'
    conteo['Categoria'] = conteo['N_Fallas'].apply(_cat)
    conteo['MTBF_dias'] = conteo.apply(
        lambda r: round(r['Dias'] / (r['N_Fallas']-1), 1) if r['N_Fallas'] > 1 else None, axis=1)
    recurrentes = conteo[conteo['N_Fallas'] > 1].copy()
    cb_rank = (df.groupby('Caja')
                 .agg(N_Fallas=('ID','count'), Strings=('String','nunique'))
                 .reset_index().sort_values('N_Fallas', ascending=False))
    return conteo, recurrentes, cb_rank

def generar_excel_fallas(planta_nombre, df, periodo="Historico"):
    AZUL='1A3A5C'; AZUL_M='2E6DA4'; ROJO='C0392B'; ROJO_C='FADBD8'
    AMAR_C='FEF9E7'; GRIS='F7F9FC'; BLC='FFFFFF'
    wb = Workbook(); ws = wb.active; ws.title='FALLAS'
    ws.sheet_view.showGridLines=False; ws.freeze_panes='A4'
    cols=[('Fecha',14),('Planta',14),('Inversor',12),('Caja',10),
          ('String',10),('Polaridad',18),('Amperios',12),('Tipo',22),('Nota',32)]
    for i,(h,w) in enumerate(cols,1):
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.merge_cells('A1:I1')
    ws['A1'].value=clean_text(f'INFORME DE FALLAS — {planta_nombre} — {periodo}')
    ws['A1'].font=_fnt(bold=True,size=13,color='FFFFFF')
    ws['A1'].fill=_fill(ROJO); ws['A1'].alignment=_aln(); ws.row_dimensions[1].height=28
    total=len(df); prom=df['Amperios'].mean() if not df.empty else 0; cortes=len(df[df['Amperios']==0])
    ws.merge_cells('A2:C2'); ws['A2'].value=f'Total: {total}'
    ws['A2'].font=_fnt(bold=True,color='FFFFFF'); ws['A2'].fill=_fill(AZUL); ws['A2'].alignment=_aln()
    ws.merge_cells('D2:F2'); ws['D2'].value=f'Promedio: {prom:.2f} A'
    ws['D2'].font=_fnt(bold=True,color='FFFFFF'); ws['D2'].fill=_fill(AZUL_M); ws['D2'].alignment=_aln()
    ws.merge_cells('G2:I2'); ws['G2'].value=f'OC (0A): {cortes}'
    ws['G2'].font=_fnt(bold=True,color='FFFFFF'); ws['G2'].fill=_fill(ROJO); ws['G2'].alignment=_aln()
    ws.row_dimensions[2].height=22
    for i,(h,_) in enumerate(cols,1): _hdr(ws.cell(3,i),h,bg=AZUL_M); ws.row_dimensions[3].height=26
    df_orig=df.copy()  # guardar antes de modificar para recurrencia
    df=df.copy(); df['Tipo']=df['Amperios'].apply(clasificar_falla_amp)
    for i,(_,r) in enumerate(df.iterrows(),4):
        tipo=str(r.get('Tipo',''))
        bg=ROJO_C if 'Corte' in tipo else (AMAR_C if 'Fatiga' in tipo else (GRIS if i%2==0 else BLC))
        fecha_str=r['Fecha'].strftime('%Y-%m-%d') if pd.notna(r.get('Fecha')) else ''
        vals=[fecha_str,r.get('Planta_Nombre',''),r.get('Inversor',''),r.get('Caja',''),
              r.get('String',''),r.get('Polaridad',''),_to_float(r.get('Amperios',0)),tipo,r.get('Nota','')]
        for j,v in enumerate(vals,1):
            _dc(ws.cell(i,j),v,bg=bg,fmt='0.00' if j==7 else None,h='left' if j==9 else 'center')
        ws.row_dimensions[i].height=18
    # ══ HOJA RECURRENCIA ════════════════════════════════════
    conteo_r, recur_r, cb_r = _calcular_recurrencia_df(df_orig)
    ws_rec = wb.create_sheet('RECURRENCIA')
    ws_rec.sheet_view.showGridLines = False
    ws_rec.freeze_panes = 'A4'

    # Título
    ws_rec.merge_cells('A1:H1')
    ws_rec['A1'].value = clean_text(f'ANALISIS DE RECURRENCIA DE FALLOS — {planta_nombre} — {periodo}')
    ws_rec['A1'].font = _fnt(bold=True, size=13, color='FFFFFF')
    ws_rec['A1'].fill = _fill(ROJO); ws_rec['A1'].alignment = _aln()
    ws_rec.row_dimensions[1].height = 28

    # KPIs fila 2
    n_total_r  = len(df)
    n_ubic_r   = len(conteo_r)
    n_recur_r  = len(recur_r)
    tasa_r     = round(n_recur_r / n_ubic_r * 100, 1) if n_ubic_r > 0 else 0
    max_r      = int(conteo_r['N_Fallas'].max()) if not conteo_r.empty else 0
    kpis_r = [
        (f'Total fallas: {n_total_r}',     'A2:B2', AZUL),
        (f'Strings afectados: {n_ubic_r}', 'C2:D2', AZUL_M),
        (f'Con recurrencia: {n_recur_r} ({tasa_r}%)', 'E2:F2', ROJO if tasa_r >= 30 else '7D5A00' if tasa_r >= 10 else '1E8449'),
        (f'Max fallas/string: {max_r}',    'G2:H2', ROJO if max_r >= 5 else AZUL),
    ]
    for txt, rng, color in kpis_r:
        ws_rec.merge_cells(rng)
        c_kpi = ws_rec[rng.split(':')[0]]
        c_kpi.value = clean_text(txt)
        c_kpi.font = _fnt(bold=True, size=10, color='FFFFFF')
        c_kpi.fill = _fill(color); c_kpi.alignment = _aln()
    ws_rec.row_dimensions[2].height = 22

    # Headers tabla
    hdrs_r = ['Ubicacion', 'N Fallas', 'Categoria', 'Primera Falla', 'Ultima Falla', 'Dias entre fallas', 'MTBF (dias)', 'Inversor']
    widths_r = [30, 10, 18, 14, 14, 18, 14, 12]
    for i, (h, w) in enumerate(zip(hdrs_r, widths_r), 1):
        ws_rec.column_dimensions[get_column_letter(i)].width = w
        _hdr(ws_rec.cell(3, i), h, bg=AZUL_M)
    ws_rec.row_dimensions[3].height = 26

    color_cat = {
        'Sin recurrencia': 'F7F9FC',
        'Recurrente (2x)': 'FEF9E7',
        'Critico (3-4x)':  'FADBD8',
        'Cronico (5+)':    'F5B7B1',
    }
    for idx, (_, row) in enumerate(conteo_r.iterrows(), 4):
        cat = str(row.get('Categoria', ''))
        bg  = color_cat.get(cat, 'F7F9FC')
        prim = row['Primera'].strftime('%d/%m/%Y') if pd.notna(row.get('Primera')) else ''
        ult  = row['Ultima'].strftime('%d/%m/%Y')  if pd.notna(row.get('Ultima'))  else ''
        vals = [str(row['Ubicacion']), int(row['N_Fallas']), clean_text(cat),
                prim, ult, int(row['Dias']),
                row['MTBF_dias'] if row['MTBF_dias'] is not None else '-',
                str(row.get('Inversor', ''))]
        fmts = [None, None, None, None, None, None, '0.0', None]
        bold_cat = cat in ('Critico (3-4x)', 'Cronico (5+)')
        for j, (v, fmt) in enumerate(zip(vals, fmts), 1):
            c = ws_rec.cell(idx, j)
            c.value = v; c.fill = _fill(bg); c.alignment = _aln(); c.border = _brd()
            c.font = _fnt(bold=bold_cat if j <= 3 else False, size=9)
            if fmt and isinstance(v, (int, float)): c.number_format = fmt
        ws_rec.row_dimensions[idx].height = 18

    # Sub-tabla: Top CBs
    start_cb = len(conteo_r) + 6
    ws_rec.merge_cells(f'A{start_cb}:D{start_cb}')
    ws_rec[f'A{start_cb}'].value = clean_text('TOP CAJAS (CB) CON MAS FALLAS')
    ws_rec[f'A{start_cb}'].font = _fnt(bold=True, size=11, color='FFFFFF')
    ws_rec[f'A{start_cb}'].fill = _fill(ROJO); ws_rec[f'A{start_cb}'].alignment = _aln()
    ws_rec.row_dimensions[start_cb].height = 22
    for j, h in enumerate(['Caja (CB)', 'N Fallas', 'Strings Afectados', '% del Total'], 1):
        _hdr(ws_rec.cell(start_cb + 1, j), h, bg=AZUL_M)
    for idx2, (_, row) in enumerate(cb_r.iterrows(), start_cb + 2):
        pct = round(int(row['N_Fallas']) / n_total_r * 100, 1) if n_total_r > 0 else 0
        for j, v in enumerate([str(row['Caja']), int(row['N_Fallas']),
                                int(row['Strings']), f'{pct}%'], 1):
            c = ws_rec.cell(idx2, j)
            c.value = v; c.fill = _fill('FADBD8' if j==2 and int(row['N_Fallas'])>3 else 'F7F9FC')
            c.alignment = _aln(); c.border = _brd(); c.font = _fnt(size=9)
        ws_rec.row_dimensions[idx2].height = 18

    out=io.BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue()

def generar_excel_mediciones(planta_nombre, df_proc, cfg=None, df_fallas=None):
    """Excel profesional 6 hojas: Portada, Mediciones, Resumen CB, Fuera de Rango, Graficos, Fallas."""
    AZUL='1A3A5C'; AZUL_M='2E6DA4'; AZUL_C='D8E8F5'; AZUL_OSC='1F5C8B'
    VERDE='1E8449'; VERDE_C='D5F5E3'; ROJO='C0392B'; ROJO_C='FADBD8'
    AMAR='F9D03F'; AMAR_C='FEF9E7'; NARANJA='E67E22'; NAR_C='FDEBD0'
    GRIS='F7F9FC'; BLC='FFFFFF'; DORADO='F4C430'

    isc_nom   = _to_float(cfg.get('Isc_STC_A',9.07)) if cfg else 9.07
    impp_nom  = _to_float(cfg.get('Impp_STC_A',8.68)) if cfg else 8.68
    pmax      = _to_float(cfg.get('Pmax_W',320)) if cfg else 320
    panels    = _to_int(cfg.get('Panels_por_String',30)) if cfg else 30
    ua        = _to_int(cfg.get('Umbral_Alerta_pct',-5)) if cfg else -5
    uc        = _to_int(cfg.get('Umbral_Critico_pct',-10)) if cfg else -10
    capacidad = str(cfg.get('Capacidad','')) if cfg else ''
    modulo    = str(cfg.get('Modulo','')) if cfg else ''
    irr       = 698
    isc_ref   = round(isc_nom * irr / 1000, 3)
    fecha_str = datetime.datetime.now().strftime('%d/%m/%Y')

    if 'String ID' not in df_proc.columns and 'String_ID' in df_proc.columns:
        df_proc = df_proc.rename(columns={'String_ID':'String ID'})

    total    = len(df_proc)
    n_norm   = len(df_proc[df_proc['Diagnostico']=='NORMAL'])
    n_aler   = len(df_proc[df_proc['Diagnostico']=='ALERTA'])
    n_crit   = len(df_proc[df_proc['Diagnostico'].isin(['CRÍTICO','OC (0A)'])])
    salud    = (n_norm/total*100) if total>0 else 0
    global_avg = df_proc['Amperios'].mean()
    df_al    = df_proc[df_proc['Diagnostico']!='NORMAL'].sort_values('Desv_CB_pct').reset_index(drop=True)

    # Resumen por CB
    cb_s = df_proc.groupby('Equipo').agg(
        N_Strings=('Amperios','count'), Imedio=('Amperios','mean'),
        Imin=('Amperios','min'), Imax=('Amperios','max'), Istd=('Amperios','std'),
        Str_Alerta=('Diagnostico', lambda x:(x=='ALERTA').sum()),
        Str_Critico=('Diagnostico', lambda x:(x.isin(['CRÍTICO','OC (0A)'])).sum()),
    ).round(3).reset_index()
    cb_s['Desv_Global_pct'] = ((cb_s['Imedio']-global_avg)/global_avg*100).round(2)
    cb_s['PR_est_pct']      = (cb_s['Imedio']/isc_ref*100).round(2)
    cb_min = cb_s.loc[cb_s['Imedio'].idxmin()]

    wb = Workbook()

    # ══ PORTADA ══════════════════════════════════════════════
    ws_p = wb.active; ws_p.title='PORTADA'
    ws_p.sheet_view.showGridLines=False
    for col,w in zip(range(1,9),[2,18,18,18,18,18,18,2]):
        ws_p.column_dimensions[get_column_letter(col)].width=w
    for r in ws_p.iter_rows(min_row=1,max_row=50,min_col=1,max_col=8):
        for c in r: c.fill=_fill(GRIS)
    for r2 in range(1,6):
        for col in range(1,9): ws_p.cell(r2,col).fill=_fill(AZUL)
    ws_p.merge_cells('B3:G3'); ws_p['B3'].value='INFORME TECNICO DE MEDICIONES DE STRINGS'
    ws_p['B3'].font=_fnt(bold=True,size=22,color=BLC); ws_p['B3'].alignment=_aln(); ws_p.row_dimensions[3].height=34
    ws_p.merge_cells('B4:G4'); ws_p['B4'].value=clean_text(f'Mundo Solar SpA      pMGD Solar — {planta_nombre} | {capacidad}')
    ws_p['B4'].font=_fnt(size=14,color='A8D1F5'); ws_p['B4'].alignment=_aln(); ws_p.row_dimensions[4].height=24
    # Barra dorada
    for col in range(2,8): ws_p.cell(6,col).fill=_fill(DORADO)

    info_rows=[
        (9,'DATOS DEL PROYECTO',None,AZUL,BLC),
        (10,'Planta:',clean_text(f'pMGD {planta_nombre}'),AZUL_C,BLC[:6]+'000000'),
        (11,'Capacidad:',capacidad,BLC,BLC[:6]+'000000'),
        (12,'Modulo FV:',clean_text(f'{modulo} / {int(pmax)} Wp'),AZUL_C,BLC[:6]+'000000'),
        (13,'Paneles/string:',str(panels),BLC,BLC[:6]+'000000'),
        (14,'Total strings:',str(total),AZUL_C,BLC[:6]+'000000'),
        (15,'Fecha medicion:',fecha_str,BLC,BLC[:6]+'000000'),
        (16,'Irradiancia:',f'~{irr} W/m2',AZUL_C,BLC[:6]+'000000'),
        (18,'PARAMETROS STC',None,AZUL,BLC),
        (19,'Isc nominal:',f'{isc_nom} A',AZUL_C,BLC[:6]+'000000'),
        (20,'Impp nominal:',f'{impp_nom} A',BLC,BLC[:6]+'000000'),
        (21,'Isc corregida:',f'{isc_ref} A',AZUL_C,BLC[:6]+'000000'),
        (22,'Umbral ALERTA:',f'< {ua}%',BLC,BLC[:6]+'000000'),
        (23,'Umbral CRITICO:',f'< {uc}%',AZUL_C,BLC[:6]+'000000'),
    ]
    for row,lb,val,bg,fg in info_rows:
        ws_p.row_dimensions[row].height=20; is_t=val is None
        ws_p.merge_cells(f'B{row}:D{row}'); c=ws_p.cell(row,2)
        c.value=lb; c.font=_fnt(bold=True,size=11 if is_t else 10,color=BLC if is_t else AZUL_OSC)
        c.fill=_fill(AZUL if is_t else bg); c.alignment=_aln(h='left'); c.border=_brd()
        if val:
            ws_p.merge_cells(f'E{row}:G{row}'); c2=ws_p.cell(row,5); c2.value=val
            c2.font=_fnt(size=10); c2.fill=_fill(bg); c2.alignment=_aln(h='left'); c2.border=_brd()

    kpi_r=26; ws_p.merge_cells(f'B{kpi_r}:G{kpi_r}')
    c=ws_p.cell(kpi_r,2); c.value='RESUMEN EJECUTIVO'
    c.font=_fnt(bold=True,size=11,color=BLC); c.fill=_fill(AZUL); c.alignment=_aln(); c.border=_brd()
    ws_p.row_dimensions[kpi_r].height=22
    kpis_port=[('Total Strings',str(total),AZUL_C,AZUL_OSC),
               ('Strings NORMAL',str(n_norm),VERDE_C,VERDE),
               ('Strings ALERTA',str(n_aler),AMAR_C,NARANJA),
               ('Strings CRITICO',str(n_crit),ROJO_C,ROJO),
               ('I Media Global',f'{global_avg:.3f} A',AZUL_C,AZUL_OSC),
               ('CB mas baja',clean_text(f"{cb_min['Equipo']} ({cb_min['Imedio']:.3f}A)"),ROJO_C,ROJO)]
    for i,(lb,val,bg,fg) in enumerate(kpis_port):
        r3=kpi_r+1+i; ws_p.row_dimensions[r3].height=20
        ws_p.merge_cells(f'B{r3}:D{r3}'); c=ws_p.cell(r3,2); c.value=lb
        c.font=_fnt(bold=True,size=10,color='444444'); c.fill=_fill(bg); c.alignment=_aln(h='left'); c.border=_brd()
        ws_p.merge_cells(f'E{r3}:G{r3}'); c2=ws_p.cell(r3,5); c2.value=val
        c2.font=_fnt(bold=True,size=11,color=fg); c2.fill=_fill(bg); c2.alignment=_aln(); c2.border=_brd()

    # ══ MEDICIONES STRINGS ═══════════════════════════════════
    ws_d=wb.create_sheet('MEDICIONES STRINGS'); ws_d.sheet_view.showGridLines=False; ws_d.freeze_panes='A4'
    cw=[5,14,10,12,14,16,16,14,14,12]
    ct=['#','Combiner Box','String','I medida (A)','Isc ref (A)','Prom. CB (A)','Desv. CB (%)','Desv. Isc (%)','Estado','P est. (W)']
    for i,(w,t) in enumerate(zip(cw,ct),1): ws_d.column_dimensions[get_column_letter(i)].width=w
    ws_d.merge_cells('A1:J1')
    ws_d['A1'].value=clean_text(f'MEDICIONES POR STRING — {planta_nombre} — {fecha_str}')
    ws_d['A1'].font=_fnt(bold=True,size=13,color=BLC); ws_d['A1'].fill=_fill(AZUL)
    ws_d['A1'].alignment=_aln(); ws_d.row_dimensions[1].height=28; ws_d.row_dimensions[3].height=30
    for col,t in enumerate(ct,1): _hdr(ws_d.cell(3,col),t,bg=AZUL_M)

    dfs=df_proc.copy()
    for idx_d,rd in dfs.reset_index(drop=True).iterrows():
        r4=idx_d+4; est=str(rd.get('Diagnostico','NORMAL')); alt=idx_d%2==0
        rb=ROJO_C if ('CRITICO' in est or 'CORTE' in est) else AMAR_C if est=='ALERTA' else (GRIS if alt else BLC)
        amp=_to_float(rd.get('Amperios',0)); prom_cb=_to_float(rd.get('Promedio_Caja',0))
        isc_r=_to_float(rd.get('Isc_ref',isc_ref)); desv_cb=_to_float(rd.get('Desv_CB_pct',0))
        desv_isc=((amp-isc_r)/isc_r*100) if isc_r>0 else 0
        pest=amp*(impp_nom/isc_nom)*panels*pmax/isc_nom if isc_nom>0 else 0
        _dc(ws_d.cell(r4,1),idx_d+1); _dc(ws_d.cell(r4,2),rd.get('Equipo',''),bg=rb)
        _dc(ws_d.cell(r4,3),rd.get('String ID',''),bg=rb)
        _dc(ws_d.cell(r4,4),amp,fmt='0.00',bold=True,bg=rb)
        _dc(ws_d.cell(r4,5),round(isc_r,3),fmt='0.000',bg=rb)
        _dc(ws_d.cell(r4,6),round(prom_cb,3),fmt='0.000',bg=rb)
        c7=ws_d.cell(r4,7); c7.value=round(desv_cb,2); c7.number_format='0.00'; c7.alignment=_aln(); c7.border=_brd()
        if 'CRITICO' in est or 'CORTE' in est: c7.font=_fnt(bold=True,color=ROJO); c7.fill=_fill(ROJO_C)
        elif est=='ALERTA': c7.font=_fnt(bold=True,color=NARANJA); c7.fill=_fill(AMAR_C)
        else: c7.font=_fnt(color=VERDE); c7.fill=_fill(rb)
        _dc(ws_d.cell(r4,8),round(desv_isc,2),fmt='0.00',bg=rb)
        c9=ws_d.cell(r4,9); c9.value=est; c9.alignment=_aln(); c9.border=_brd()
        if 'CRITICO' in est or 'CORTE' in est: c9.font=_fnt(bold=True,color=BLC); c9.fill=_fill(ROJO)
        elif est=='ALERTA': c9.font=_fnt(bold=True,color='7D4F00'); c9.fill=_fill(AMAR)
        else: c9.font=_fnt(bold=True,color=BLC); c9.fill=_fill(VERDE)
        _dc(ws_d.cell(r4,10),round(pest,1),fmt='#,##0.0',bg=rb)
        ws_d.row_dimensions[r4].height=18

    lr=len(dfs)+4
    for col in range(1,11): c=ws_d.cell(lr,col); c.font=_fnt(bold=True,color=BLC); c.fill=_fill(AZUL_OSC); c.alignment=_aln(); c.border=_brd()
    ws_d.cell(lr,1).value='PROMEDIO'
    ws_d.cell(lr,4).value=f'=AVERAGE(D4:D{lr-1})'; ws_d.cell(lr,4).number_format='0.000'
    ws_d.cell(lr,10).value=f'=SUM(J4:J{lr-1})'; ws_d.cell(lr,10).number_format='#,##0'

    # ══ RESUMEN POR CB ════════════════════════════════════════
    ws_cb=wb.create_sheet('RESUMEN POR CB'); ws_cb.sheet_view.showGridLines=False; ws_cb.freeze_panes='A4'
    cw2=[5,14,10,12,10,10,10,16,12,12,12]
    ct2=['#','Combiner Box','N Strings','I media (A)','I min (A)','I max (A)','Std Dev','Desv. Global (%)','PR est. (%)','Str. Alerta','Str. Critico']
    for i,(w,t) in enumerate(zip(cw2,ct2),1): ws_cb.column_dimensions[get_column_letter(i)].width=w
    ws_cb.merge_cells('A1:K1')
    ws_cb['A1'].value=clean_text(f'RESUMEN POR CB — {planta_nombre} — {fecha_str}')
    ws_cb['A1'].font=_fnt(bold=True,size=13,color=BLC); ws_cb['A1'].fill=_fill(AZUL)
    ws_cb['A1'].alignment=_aln(); ws_cb.row_dimensions[1].height=28; ws_cb.row_dimensions[3].height=30
    for col,t in enumerate(ct2,1): _hdr(ws_cb.cell(3,col),t,bg=AZUL_M)
    for idx_c,rd in cb_s.reset_index(drop=True).iterrows():
        r5=idx_c+4; alt=idx_c%2==0; bg=GRIS if alt else BLC
        desv=rd['Desv_Global_pct']; pr=rd['PR_est_pct']
        al=int(rd['Str_Alerta']); cr=int(rd['Str_Critico'])
        _dc(ws_cb.cell(r5,1),idx_c+1); _dc(ws_cb.cell(r5,2),rd['Equipo'],bold=True,color=AZUL_OSC,bg=bg)
        _dc(ws_cb.cell(r5,3),int(rd['N_Strings']),bg=bg); _dc(ws_cb.cell(r5,4),round(rd['Imedio'],3),fmt='0.000',bold=True,bg=bg)
        _dc(ws_cb.cell(r5,5),round(rd['Imin'],2),fmt='0.00',bg=bg); _dc(ws_cb.cell(r5,6),round(rd['Imax'],2),fmt='0.00',bg=bg)
        _dc(ws_cb.cell(r5,7),round(rd['Istd'],3),fmt='0.000',bg=bg)
        c8=ws_cb.cell(r5,8); c8.value=round(desv,2); c8.number_format='0.00'; c8.alignment=_aln(); c8.border=_brd()
        if desv<-5: c8.font=_fnt(bold=True,color=ROJO); c8.fill=_fill(ROJO_C)
        elif desv<0: c8.font=_fnt(color=NARANJA); c8.fill=_fill(bg)
        else: c8.font=_fnt(color=VERDE); c8.fill=_fill(bg)
        c9=ws_cb.cell(r5,9); c9.value=round(pr,2); c9.number_format='0.00'; c9.alignment=_aln(); c9.border=_brd()
        c9.fill=_fill(VERDE_C if pr>=100 else AMAR_C if pr>=95 else ROJO_C); c9.font=_fnt(bold=True)
        c10=ws_cb.cell(r5,10); c10.value=al; c10.alignment=_aln(); c10.border=_brd()
        c10.fill=_fill(AMAR_C if al>0 else bg); c10.font=_fnt(bold=al>0,color=NARANJA if al>0 else '000000')
        c11=ws_cb.cell(r5,11); c11.value=cr; c11.alignment=_aln(); c11.border=_brd()
        c11.fill=_fill(ROJO_C if cr>0 else bg); c11.font=_fnt(bold=cr>0,color=ROJO if cr>0 else '000000')
        ws_cb.row_dimensions[r5].height=20

    # ══ STRINGS FUERA DE RANGO ═══════════════════════════════
    ws_al=wb.create_sheet('STRINGS FUERA DE RANGO'); ws_al.sheet_view.showGridLines=False
    cw3=[5,14,10,12,14,16,12,30,25]
    ct3=['#','Combiner Box','String','I medida (A)','Prom. CB (A)','Desv. CB (%)','Estado','Posible Causa','Accion Recomendada']
    for i,(w,t) in enumerate(zip(cw3,ct3),1): ws_al.column_dimensions[get_column_letter(i)].width=w
    ws_al.merge_cells('A1:I1')
    ws_al['A1'].value=clean_text(f'STRINGS FUERA DE RANGO — {len(df_al)} identificados — {planta_nombre}')
    ws_al['A1'].font=_fnt(bold=True,size=13,color=BLC); ws_al['A1'].fill=_fill(ROJO)
    ws_al['A1'].alignment=_aln(); ws_al.row_dimensions[1].height=28; ws_al.row_dimensions[3].height=30
    for col,t in enumerate(ct3,1): _hdr(ws_al.cell(3,col),t,bg=AZUL)
    def causa_xl(d):
        if d<=uc: return 'Modulo(s) defectuoso(s), conector MC4 danado o bypass activado'
        if d<=-7: return 'Modulo degradado, suciedad intensa o sombra parcial'
        return 'Suciedad, sombra leve o degradacion inicial'
    def accion_xl(d):
        if d<=uc: return 'Inspeccion urgente + termografia + curva I-V'
        if d<=-7: return 'Inspeccion visual + limpieza + revision conectores'
        return 'Monitorear + limpieza preventiva'
    for idx_a,rd in df_al.iterrows():
        r6=idx_a+4; desv=_to_float(rd.get('Desv_CB_pct',0)); est=str(rd.get('Diagnostico',''))
        bg=ROJO_C if ('CRITICO' in est or 'CORTE' in est) else AMAR_C
        _dc(ws_al.cell(r6,1),idx_a+1); _dc(ws_al.cell(r6,2),rd.get('Equipo',''),bold=True,bg=bg)
        _dc(ws_al.cell(r6,3),rd.get('String ID',''),bg=bg)
        _dc(ws_al.cell(r6,4),_to_float(rd.get('Amperios',0)),fmt='0.00',bold=True,bg=bg)
        _dc(ws_al.cell(r6,5),round(_to_float(rd.get('Promedio_Caja',0)),3),fmt='0.000',bg=bg)
        c6=ws_al.cell(r6,6); c6.value=round(desv,2); c6.number_format='0.00'; c6.alignment=_aln(); c6.border=_brd()
        c6.font=_fnt(bold=True,color=ROJO if 'CRITICO' in est or 'CORTE' in est else NARANJA); c6.fill=_fill(bg)
        c7=ws_al.cell(r6,7); c7.value=est; c7.alignment=_aln(); c7.border=_brd()
        if 'CRITICO' in est or 'CORTE' in est: c7.font=_fnt(bold=True,color=BLC); c7.fill=_fill(ROJO)
        else: c7.font=_fnt(bold=True,color='7D4F00'); c7.fill=_fill(AMAR)
        for col_n,txt in [(8,causa_xl(desv)),(9,accion_xl(desv))]:
            c=ws_al.cell(r6,col_n); c.value=clean_text(txt); c.fill=_fill(bg); c.border=_brd()
            c.alignment=Alignment(horizontal='left',vertical='center',wrap_text=True); c.font=_fnt(size=9)
        ws_al.row_dimensions[r6].height=28

    # ══ GRÁFICOS ═════════════════════════════════════════════
    from openpyxl.chart import BarChart, Reference
    ws_g=wb.create_sheet('GRAFICOS'); ws_g.sheet_view.showGridLines=False
    ws_g.merge_cells('A1:P1'); ws_g['A1'].value=clean_text(f'GRAFICOS — {planta_nombre} {capacidad}')
    ws_g['A1'].font=_fnt(bold=True,size=13,color=BLC); ws_g['A1'].fill=_fill(AZUL)
    ws_g['A1'].alignment=_aln(); ws_g.row_dimensions[1].height=28
    for i in range(1,7): ws_g.column_dimensions[get_column_letter(i)].width=14
    g_hdrs=['CB ID','I media (A)','I min (A)','I max (A)','PR est. (%)','Alertas+Criticos']
    for col,h in enumerate(g_hdrs,1): _hdr(ws_g.cell(3,col),h,bg=AZUL_M,size=9)
    for i,rd in cb_s.reset_index(drop=True).iterrows():
        r7=i+4; alt=i%2==0; bg=GRIS if alt else BLC
        vals=[rd['Equipo'],round(rd['Imedio'],3),round(rd['Imin'],2),round(rd['Imax'],2),
              round(rd['PR_est_pct'],2),int(rd['Str_Alerta'])+int(rd['Str_Critico'])]
        fmts=[None,'0.000','0.00','0.00','0.00',None]
        for col,(v,fmt) in enumerate(zip(vals,fmts),1):
            c=ws_g.cell(r7,col); c.value=v; c.fill=_fill(bg); c.alignment=_aln(); c.border=_brd(); c.font=_fnt(size=9)
            if fmt: c.number_format=fmt
    last_rg=len(cb_s)+4
    for title,col_g,y_min,y_max,anchor in [
        ('Corriente Media por CB',2,4.0,9.5,'A20'),
        ('PR Estimado por CB (%)',5,75,115,'L20'),
        ('Strings Alerta+Critico',6,None,None,'A42')
    ]:
        ch=BarChart(); ch.type='col'; ch.title=title; ch.style=10; ch.width=26; ch.height=14
        data=Reference(ws_g,min_col=col_g,max_col=col_g,min_row=3,max_row=last_rg)
        cats=Reference(ws_g,min_col=1,min_row=4,max_row=last_rg)
        ch.add_data(data,titles_from_data=True); ch.set_categories(cats)
        if y_min: ch.y_axis.scaling.min=y_min
        if y_max: ch.y_axis.scaling.max=y_max
        ws_g.add_chart(ch,anchor)

    # ══ FALLAS ═══════════════════════════════════════════════
    ws_fal=wb.create_sheet('FALLAS')
    if df_fallas is not None and not df_fallas.empty:
        ws_fal.sheet_view.showGridLines=False; ws_fal.freeze_panes='A4'
        cols_fal=[('Fecha',14),('Inversor',12),('Caja',10),('String',10),
                  ('Polaridad',18),('Amperios',12),('Tipo',22),('Nota',32)]
        for i,(h,w) in enumerate(cols_fal,1): ws_fal.column_dimensions[get_column_letter(i)].width=w
        ws_fal.merge_cells('A1:H1')
        ws_fal['A1'].value=clean_text(f'FALLAS REGISTRADAS — {planta_nombre}')
        ws_fal['A1'].font=_fnt(bold=True,size=13,color=BLC)
        ws_fal['A1'].fill=_fill(ROJO); ws_fal['A1'].alignment=_aln(); ws_fal.row_dimensions[1].height=28
        tf=len(df_fallas); pf=df_fallas['Amperios'].mean(); cf=len(df_fallas[df_fallas['Amperios']==0])
        ws_fal.merge_cells('A2:C2'); ws_fal['A2'].value=f'Total: {tf}'
        ws_fal['A2'].font=_fnt(bold=True,color=BLC); ws_fal['A2'].fill=_fill(AZUL); ws_fal['A2'].alignment=_aln()
        ws_fal.merge_cells('D2:F2'); ws_fal['D2'].value=f'Promedio: {pf:.2f} A'
        ws_fal['D2'].font=_fnt(bold=True,color=BLC); ws_fal['D2'].fill=_fill(AZUL_M); ws_fal['D2'].alignment=_aln()
        ws_fal.merge_cells('G2:H2'); ws_fal['G2'].value=f'OC (0A): {cf}'
        ws_fal['G2'].font=_fnt(bold=True,color=BLC); ws_fal['G2'].fill=_fill(ROJO); ws_fal['G2'].alignment=_aln()
        ws_fal.row_dimensions[2].height=22
        for i,(h,_) in enumerate(cols_fal,1): _hdr(ws_fal.cell(3,i),h,bg=AZUL_M); ws_fal.row_dimensions[3].height=26
        df_fal2=df_fallas.copy(); df_fal2['Tipo']=df_fal2['Amperios'].apply(clasificar_falla_amp)
        for i,(_,r) in enumerate(df_fal2.iterrows(),4):
            tipo=str(r.get('Tipo','')); alt=i%2==0
            bg=ROJO_C if 'Corte' in tipo else (AMAR_C if 'Fatiga' in tipo else (GRIS if alt else BLC))
            fecha_s=r['Fecha'].strftime('%Y-%m-%d') if pd.notna(r.get('Fecha')) else ''
            vals=[fecha_s,r.get('Inversor',''),r.get('Caja',''),r.get('String',''),
                  r.get('Polaridad',''),_to_float(r.get('Amperios',0)),tipo,r.get('Nota','')]
            for j,v in enumerate(vals,1):
                _dc(ws_fal.cell(i,j),v,bg=bg,fmt='0.00' if j==6 else None,h='left' if j==8 else 'center')
            ws_fal.row_dimensions[i].height=18
    else:
        ws_fal['A1'].value='Sin fallas registradas para esta planta.'
        ws_fal['A1'].font=_fnt(size=11); ws_fal['A1'].alignment=_aln()

    out=io.BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue()


